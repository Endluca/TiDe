from __future__ import annotations

import hashlib
import json
import math
import re
from copy import deepcopy
from dataclasses import asdict, dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from sqlalchemy import Engine, case, func, select
from sqlalchemy.orm import Session

from .config_models import (
    DEFAULT_CONFIG_PAYLOADS,
    SCORE_POLICY_V2_PAYLOAD,
    ConfigKey,
    ConfigStatus,
    ConfigVersionRecord,
    ScoreGraduationConfig,
)
from .database import engine as default_engine
from .database import session_scope
from .db_models import (
    DataImportBatchRecord,
    ScoreAccountRecord,
    ScoreEntryRecord,
    TaskAssignmentRecord,
    TeacherMetricSnapshotRecord,
    TeacherRecord,
)
from .fixed_growth_baseline import (
    FixedGrowthBaselineError,
    ensure_fixed_growth_assignments,
)


SOURCE_SHEET = "境外教师明细"
SOURCE_SYSTEM = "MANUAL_XLSX:OVERSEAS_NEW_TEACHER_30D_WIDE"
SCORE_RULE_VERSION = "new_teacher_30d_20260723_v6"
SCORE_POLICY_SNAPSHOT = ScoreGraduationConfig.model_validate(
    DEFAULT_CONFIG_PAYLOADS[ConfigKey.SCORE_GRADUATION]
).model_dump(mode="json")
SCORE_POLICY_V2_SNAPSHOT = ScoreGraduationConfig.model_validate(
    SCORE_POLICY_V2_PAYLOAD
).model_dump(mode="json")
CAPACITY_MILESTONE_ID = "CAPACITY_PEAK_SLOT_40"
CAPACITY_MILESTONE_REASON_CODE = "CAPACITY_MILESTONE_ACHIEVED"
CAPACITY_MILESTONE_SETTLEMENT_MODE = "FIRST_ACHIEVEMENT_LOCKED"
CAPACITY_MILESTONE_POLICY_VERSIONS = frozenset({"v4", "v5", "v6"})
DIRECT_EXTERNAL_SCALE_POLICY_VERSIONS = frozenset({"v3", "v4", "v5", "v6"})


def score_policy_sha256(policy: dict[str, Any]) -> str:
    encoded = json.dumps(
        policy,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


SCORE_POLICY_SHA256 = score_policy_sha256(SCORE_POLICY_SNAPSHOT)

# This is intentionally an exact contract. A renamed, omitted, reordered, or
# additional source column must be reviewed before it can enter production.
EXPECTED_HEADERS: tuple[str, ...] = (
    "tchr_id",
    "real_name",
    "tchr_group",
    "tchr_group_desc",
    "center_type_id",
    "center_type_desc",
    "bu",
    "based_type",
    "status",
    "status_on_date",
    "status_off_date",
    "last_on_date",
    "job_days",
    "job_month",
    "is_ft_hbt",
    "is_fte",
    "teach_area_type",
    "tchr_score",
    "onboard_date",
    "onboard_30d_end_date",
    "first_open_slot_dt",
    "first_booked_dt",
    "first_completed_dt",
    "total_booked_cnt",
    "peak_booked_cnt",
    "total_completed_cnt",
    "peak_completed_cnt",
    "absent_cnt",
    "late_cnt",
    "early_cnt",
    "anomaly_cnt",
    "perfect_cnt",
    "no_notice_cnt",
    "first_completed_student_cnt",
    "completed_again_student_15d_cnt",
    "feedback_total_eval_cnt",
    "feedback_praise_cnt",
    "feedback_negative_cnt",
    "feedback_complaint_cnt",
    "feedback_valid_complaint_cnt",
    "feedback_favorite_cnt",
    "feedback_block_cnt",
    "total_slot_cnt",
    "reg_slot_cnt",
    "peak_slot_cnt",
    "slot_days",
    "peak_slot_days",
    "reliability_absent_rate",
    "reliability_late_rate",
    "reliability_early_leave_rate",
    "reliability_late_early_rate",
    "feedback_praise_rate",
    "feedback_negative_rate",
    "feedback_complaint_rate",
    "feedback_rebook_rate",
    "feedback_favorite_rate",
    "feedback_block_rate",
    "feedback_eval_rate",
    "capacity_avg_completed_per_day",
    "capacity_peak_slot_rate",
    "capacity_key_slot_day_rate",
)


class ImportValidationError(ValueError):
    """Raised before persistence when a workbook violates the source contract."""


@dataclass(frozen=True)
class TeacherMetricImportResult:
    batch_id: str
    source_uri: str
    source_sha256: str
    source_sheet: str
    snapshot_label: str
    row_count: int
    column_count: int
    data_mode: str
    idempotent_reimport: bool
    perfect_vs_on_time_difference_count: int

    def as_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class ClassQualityRecalculationResult:
    """Outcome of an explicit current-snapshot classroom-quality rebuild."""

    policy_version: str
    score_policy_sha256: str
    points_per_unit: float
    teachers_scanned: int
    current_snapshots_found: int
    snapshots_updated: int
    teacher_payloads_updated: int
    score_accounts_created: int
    score_accounts_updated: int
    teachers_without_current_snapshot: int
    dry_run: bool

    def as_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class _ValidatedRow:
    row_number: int
    teacher_id: str
    raw_payload: dict[str, Any]


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        normalized = value if value.tzinfo else value.replace(tzinfo=timezone.utc)
        return normalized.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        raise ImportValidationError("source contains a non-finite number")
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _teacher_id(value: Any, *, row_number: int) -> str:
    if value is None or value == "":
        raise ImportValidationError(f"row {row_number}: tchr_id is required")
    if isinstance(value, bool):
        raise ImportValidationError(f"row {row_number}: boolean tchr_id is invalid")
    if isinstance(value, float):
        if not value.is_integer():
            raise ImportValidationError(f"row {row_number}: tchr_id must be an integer identifier")
        value = int(value)
    normalized = str(value).strip()
    if not normalized:
        raise ImportValidationError(f"row {row_number}: tchr_id is required")
    if len(normalized) > 64:
        raise ImportValidationError(f"row {row_number}: tchr_id exceeds 64 characters")
    return normalized


def _as_nonnegative_int(raw: dict[str, Any], field: str, *, row_number: int) -> int:
    value = raw.get(field)
    if value in (None, ""):
        return 0
    if isinstance(value, bool):
        converted = int(value)
    elif isinstance(value, int):
        converted = value
    elif isinstance(value, float) and value.is_integer():
        converted = int(value)
    else:
        try:
            converted = int(str(value).strip())
        except (TypeError, ValueError) as exc:
            raise ImportValidationError(
                f"row {row_number}: {field} must be an integer, got {value!r}"
            ) from exc
    if converted < 0:
        raise ImportValidationError(f"row {row_number}: {field} cannot be negative")
    return converted


def _as_optional_date(raw: dict[str, Any], field: str, *, row_number: int) -> date | None:
    value = raw.get(field)
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    normalized = str(value).strip()
    if not normalized:
        return None
    try:
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", normalized):
            return date.fromisoformat(normalized)
        if not re.match(r"^\d{4}-\d{2}-\d{2}[T ]", normalized):
            raise ValueError
        return datetime.fromisoformat(normalized.replace("Z", "+00:00")).date()
    except ValueError as exc:
        raise ImportValidationError(
            f"row {row_number}: {field} must be an ISO/Excel date, got {value!r}"
        ) from exc


def _as_optional_boolean(
    raw: dict[str, Any], field: str, *, row_number: int
) -> bool | None:
    """Preserve missing upstream evidence and accept only controlled booleans."""

    value = raw.get(field)
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int) and value in (0, 1):
        return bool(value)
    if isinstance(value, float) and value in (0.0, 1.0):
        return bool(int(value))
    if isinstance(value, str) and value.strip() in {"0", "1"}:
        return value.strip() == "1"
    raise ImportValidationError(
        f"row {row_number}: {field} must be an explicit boolean or controlled 0/1, "
        f"got {value!r}"
    )


def _avatar(name: str, teacher_id: str) -> str:
    parts = [part for part in name.replace("-", " ").split() if part]
    if parts:
        return "".join(part[0].upper() for part in parts[:2])
    return teacher_id[:2].upper()


def _public_score(
    raw_total_score: float,
    score_policy: dict[str, Any] | None = None,
) -> float:
    policy = score_policy or SCORE_POLICY_SNAPSHOT
    thresholds = policy["thresholds"]
    if policy.get("policy_version") in DIRECT_EXTERNAL_SCALE_POLICY_VERSIONS:
        return round(min(raw_total_score, thresholds["gold_external_score"]), 2)
    if raw_total_score < thresholds["graduation_raw_score"]:
        return round(
            raw_total_score
            * thresholds["graduation_external_score"]
            / thresholds["graduation_raw_score"],
            2,
        )
    if raw_total_score < thresholds["gold_raw_score"]:
        return round(
            thresholds["graduation_external_score"]
            + (raw_total_score - thresholds["graduation_raw_score"])
            * (thresholds["gold_external_score"] - thresholds["graduation_external_score"])
            / (thresholds["gold_raw_score"] - thresholds["graduation_raw_score"]),
            2,
        )
    return round(float(thresholds["gold_external_score"]), 2)


def _provenance(
    *,
    source_mode: str,
    source_field: str | None,
    batch_id: str,
    note: str,
    source_fields: list[str] | None = None,
) -> dict[str, Any]:
    item: dict[str, Any] = {
        "source_mode": source_mode,
        "source_field": source_field,
        "batch_id": batch_id,
        "note": note,
    }
    if source_fields is not None:
        item["source_fields"] = source_fields
    return item


def _teacher_profile_projection(
    raw: dict[str, Any],
    *,
    batch_id: str,
    row_number: int,
) -> dict[str, Any]:
    """Project source profile facts without converting missing evidence to false."""

    first_booked_date = _as_optional_date(
        raw, "first_booked_dt", row_number=row_number
    )
    is_cpl_tesol = _as_optional_boolean(
        raw, "is_cpl_tesol", row_number=row_number
    )
    is_self_introduce = _as_optional_boolean(
        raw, "is_self_introduce", row_number=row_number
    )
    return {
        "first_booked_date": first_booked_date,
        "is_cpl_tesol": is_cpl_tesol,
        "is_self_introduce": is_self_introduce,
        "payload": {
            "first_booked_date": (
                first_booked_date.isoformat() if first_booked_date else None
            ),
            "is_cpl_tesol": is_cpl_tesol,
            "is_self_introduce": is_self_introduce,
        },
        "provenance": {
            "first_booked_date": {
                "source_mode": "REAL" if first_booked_date else "SOURCE_MISSING",
                "source_field": "first_booked_dt",
                "batch_id": batch_id,
                "note": (
                    "Date-level first booked lesson value from the validated teacher snapshot; "
                    "it does not prove that the lesson was completed."
                    if first_booked_date
                    else "The source workbook has no first booked lesson date for this teacher."
                ),
            },
            "is_cpl_tesol": {
                "source_mode": "REAL" if is_cpl_tesol is not None else "SOURCE_MISSING",
                "source_field": "is_cpl_tesol",
                "batch_id": batch_id,
                "note": (
                    "Explicit upstream TESOL-completion evidence; it is not G01 completion."
                    if is_cpl_tesol is not None
                    else "The current 61-column workbook does not provide TESOL completion."
                ),
            },
            "is_self_introduce": {
                "source_mode": (
                    "REAL" if is_self_introduce is not None else "SOURCE_MISSING"
                ),
                "source_field": "is_self_introduce",
                "batch_id": batch_id,
                "note": (
                    "Explicit upstream self-introduction evidence; it is not G01 completion."
                    if is_self_introduce is not None
                    else "The current 61-column workbook does not provide self-introduction completion."
                ),
            },
        },
    }


def _capacity_milestone_idempotency_key(teacher_id: str) -> str:
    return f"CAPACITY_MILESTONE:{CAPACITY_MILESTONE_ID}:{teacher_id}"


def _capacity_milestone_score_entry_id(teacher_id: str) -> str:
    digest = hashlib.sha256(
        _capacity_milestone_idempotency_key(teacher_id).encode("utf-8")
    ).hexdigest()[:32]
    return f"CAPM-{digest}"


def _capacity_milestone_entry_payload(
    *,
    teacher_id: str,
    camp_enrollment_id: str,
    batch_id: str,
    snapshot_id: str,
    peak_slot_cnt: int,
    occurred_at: datetime,
) -> dict[str, Any]:
    entry_id = _capacity_milestone_score_entry_id(teacher_id)
    return {
        "score_entry_id": entry_id,
        "camp_enrollment_id": camp_enrollment_id,
        "lesson_id": None,
        "teacher_id": teacher_id,
        "dimension": "CAPACITY",
        "entry_type": "MILESTONE_ACHIEVEMENT",
        "delta_score": 10.0,
        "reason_code": CAPACITY_MILESTONE_REASON_CODE,
        "evidence_status": "CONFIRMED",
        "score_rule_version": SCORE_RULE_VERSION,
        "occurred_at": occurred_at.isoformat(),
        "reversal_of_score_entry_id": None,
        "idempotency_key": _capacity_milestone_idempotency_key(teacher_id),
        "milestone_id": CAPACITY_MILESTONE_ID,
        "metric": "peak_slot_cnt",
        "operator": "GTE",
        "threshold": 40,
        "observed_value": peak_slot_cnt,
        "source_batch_id": batch_id,
        "source_snapshot_id": snapshot_id,
        "source_mode": "DERIVED_REAL",
        "settlement_mode": CAPACITY_MILESTONE_SETTLEMENT_MODE,
    }


def _find_capacity_milestone_entry(
    session: Session,
    teacher_id: str,
) -> ScoreEntryRecord | None:
    existing = session.scalar(
        select(ScoreEntryRecord).where(
            ScoreEntryRecord.idempotency_key
            == _capacity_milestone_idempotency_key(teacher_id)
        )
    )
    if existing is not None:
        return existing
    # This secondary lookup protects a database that already contains a v4
    # milestone row created before the deterministic idempotency key was frozen.
    return session.scalar(
        select(ScoreEntryRecord)
        .where(
            ScoreEntryRecord.teacher_id == teacher_id,
            ScoreEntryRecord.dimension == "CAPACITY",
            ScoreEntryRecord.reason_code == CAPACITY_MILESTONE_REASON_CODE,
        )
        .order_by(ScoreEntryRecord.recorded_at, ScoreEntryRecord.score_entry_id)
    )


def _append_score_entry_payload(
    teacher_payload: dict[str, Any],
    entry_payload: dict[str, Any],
) -> None:
    score_entries = list(teacher_payload.get("score_entries") or [])
    entry_id = entry_payload["score_entry_id"]
    if not any(
        isinstance(item, dict) and item.get("score_entry_id") == entry_id
        for item in score_entries
    ):
        score_entries.append(deepcopy(entry_payload))
    teacher_payload["score_entries"] = score_entries


def _score_entry_payload_from_record(entry: ScoreEntryRecord) -> dict[str, Any]:
    payload = deepcopy(entry.payload or {})
    payload.update(
        score_entry_id=entry.score_entry_id,
        camp_enrollment_id=entry.camp_enrollment_id,
        lesson_id=entry.lesson_id,
        teacher_id=entry.teacher_id,
        dimension=entry.dimension,
        entry_type=entry.entry_type,
        delta_score=float(entry.delta_score),
        reason_code=entry.reason_code,
        evidence_status=entry.evidence_status,
        score_rule_version=entry.score_rule_version,
        occurred_at=entry.occurred_at.isoformat() if entry.occurred_at else None,
        reversal_of_score_entry_id=entry.reversal_of_score_entry_id,
        idempotency_key=entry.idempotency_key,
    )
    return payload


def _ensure_capacity_milestone_entry(
    session: Session,
    *,
    teacher: TeacherRecord,
    batch_id: str,
    snapshot_id: str,
    peak_slot_cnt: int,
    occurred_at: datetime,
) -> tuple[ScoreEntryRecord, dict[str, Any], bool]:
    existing = _find_capacity_milestone_entry(session, teacher.teacher_id)
    if existing is not None:
        return existing, _score_entry_payload_from_record(existing), False

    payload = _capacity_milestone_entry_payload(
        teacher_id=teacher.teacher_id,
        camp_enrollment_id=teacher.camp_enrollment_id,
        batch_id=batch_id,
        snapshot_id=snapshot_id,
        peak_slot_cnt=peak_slot_cnt,
        occurred_at=occurred_at,
    )
    entry = ScoreEntryRecord(
        score_entry_id=payload["score_entry_id"],
        camp_enrollment_id=teacher.camp_enrollment_id,
        lesson_id=None,
        teacher_id=teacher.teacher_id,
        dimension="CAPACITY",
        entry_type="MILESTONE_ACHIEVEMENT",
        delta_score=10.0,
        reason_code=CAPACITY_MILESTONE_REASON_CODE,
        evidence_status="CONFIRMED",
        score_rule_version=SCORE_RULE_VERSION,
        occurred_at=occurred_at,
        recorded_at=occurred_at,
        reversal_of_score_entry_id=None,
        idempotency_key=payload["idempotency_key"],
        payload=deepcopy(payload),
    )
    session.add(entry)
    return entry, payload, True


def _metric_projection(
    raw: dict[str, Any],
    batch_id: str,
    *,
    row_number: int,
    score_policy: dict[str, Any] | None = None,
    prior_capacity_milestone_achieved: bool = False,
) -> dict[str, Any]:
    effective_policy = score_policy or SCORE_POLICY_SNAPSHOT
    policy_version = str(effective_policy.get("policy_version") or "v3")
    total_completed = _as_nonnegative_int(raw, "total_completed_cnt", row_number=row_number)
    peak_completed = _as_nonnegative_int(raw, "peak_completed_cnt", row_number=row_number)
    peak_slot_value_missing = raw.get("peak_slot_cnt") in (None, "")
    peak_slot_count = _as_nonnegative_int(raw, "peak_slot_cnt", row_number=row_number)
    perfect_value_missing = raw.get("perfect_cnt") in (None, "")
    perfect = _as_nonnegative_int(raw, "perfect_cnt", row_number=row_number)
    praise = _as_nonnegative_int(raw, "feedback_praise_cnt", row_number=row_number)
    favorite = _as_nonnegative_int(raw, "feedback_favorite_cnt", row_number=row_number)
    rebook = _as_nonnegative_int(
        raw, "completed_again_student_15d_cnt", row_number=row_number
    )
    late = _as_nonnegative_int(raw, "late_cnt", row_number=row_number)
    early = _as_nonnegative_int(raw, "early_cnt", row_number=row_number)
    _as_nonnegative_int(raw, "absent_cnt", row_number=row_number)

    # Punctual completion remains a separate reliability input. Classroom
    # quality uses the source's reviewed perfect-completion count under v5+.
    on_time_completed = max(total_completed - late - early, 0)
    scoring_items = effective_policy["scoring_items"]
    if policy_version in CAPACITY_MILESTONE_POLICY_VERSIONS:
        capacity_rule = scoring_items["capacity"]
        current_capacity_threshold_met = peak_slot_count >= int(capacity_rule["threshold"])
        capacity_milestone_achieved = bool(
            prior_capacity_milestone_achieved or current_capacity_threshold_met
        )
        capacity_score_input = (
            float(capacity_rule["score_value"]) if capacity_milestone_achieved else 0.0
        )
    else:
        # Frozen v2/v3 snapshots used a temporary fixed 10-point assumption.
        current_capacity_threshold_met = False
        capacity_milestone_achieved = False
        capacity_score_input = 10.0

    metric_inputs = {
        "total_completed_cnt": total_completed,
        "peak_completed_cnt": peak_completed,
        "perfect_cnt": perfect,
        "on_time_completed_cnt": on_time_completed,
        "feedback_praise_cnt": praise,
        "feedback_favorite_cnt": favorite,
        "completed_again_student_15d_cnt": rebook,
        "late_cnt": late,
        "early_cnt": early,
        # The source absent_cnt mixes physical absence, leave and system
        # cancellation. Keep it in raw_payload, but do not use it as confirmed
        # real-absence evidence.
        "real_absent_cnt": 0,
        "severe_redline_event": False,
        # L0 complaints are derived from the matched lesson/complaint table,
        # not from this teacher-wide source.
        "l0_complaint_cnt": 0,
        "capacity_score": capacity_score_input,
        # Fixed-task points are settled only from completed G01-G10 assignments.
        "new_teacher_task_score": 0,
        "mandatory_task_assignment_count": 0,
        "mandatory_task_completed_count": 0,
        "mandatory_task_expected_count": 10,
        # Historical compatibility column. The current classroom-quality
        # formula reads perfect_cnt directly and does not use this rate.
        "class_quality_no_issue_rate": 0.0,
    }
    if policy_version in CAPACITY_MILESTONE_POLICY_VERSIONS:
        metric_inputs.update(
            peak_slot_cnt=peak_slot_count,
            capacity_milestone_id=capacity_rule["milestone_id"],
            capacity_milestone_achieved=capacity_milestone_achieved,
            capacity_milestone_currently_meets_threshold=current_capacity_threshold_met,
            capacity_milestone_settlement_mode=capacity_rule["settlement_mode"],
        )
    metric_provenance = {
        field: _provenance(
            source_mode="REAL",
            source_field=field,
            batch_id=batch_id,
            note="Direct value from the validated 61-column teacher snapshot.",
        )
        for field in (
            "total_completed_cnt",
            "peak_completed_cnt",
            "perfect_cnt",
            "feedback_praise_cnt",
            "feedback_favorite_cnt",
            "completed_again_student_15d_cnt",
            "late_cnt",
            "early_cnt",
        )
    }
    metric_provenance["on_time_completed_cnt"] = _provenance(
        source_mode="DERIVED_REAL",
        source_field="total_completed_cnt,late_cnt,early_cnt",
        source_fields=["total_completed_cnt", "late_cnt", "early_cnt"],
        batch_id=batch_id,
        note=(
            f"{policy_version} reproducible backtest convention: max(total_completed_cnt - late_cnt - "
            "early_cnt, 0); perfect_cnt is retained separately for audit."
        ),
    )
    metric_provenance["perfect_cnt"] = _provenance(
        source_mode="SOURCE_MISSING" if perfect_value_missing else "REAL",
        source_field="perfect_cnt",
        batch_id=batch_id,
        note=(
            "Source perfect_cnt is empty; it is normalized to 0 but cannot be "
            "presented as an observed classroom-quality count."
            if perfect_value_missing
            else "Direct perfect-completion count from the validated teacher snapshot."
        ),
    )
    metric_provenance["real_absent_cnt"] = _provenance(
        source_mode="SOURCE_MISSING",
        source_field=None,
        source_fields=["absent_cnt"],
        batch_id=batch_id,
        note=(
            "Confirmed physical-absence evidence is unavailable. The source absent_cnt "
            "mixes leave and system-cancel cases, so scoring uses zero and retains "
            "absent_cnt only in the lossless source row."
        ),
    )
    metric_provenance["severe_redline_event"] = _provenance(
        source_mode="SOURCE_MISSING",
        source_field=None,
        batch_id=batch_id,
        note=(
            "The delivered workbook has no severe-redline-event field. The stored false "
            "is only a non-null placeholder; qualification remains blocked by missing "
            "mandatory-task evidence."
        ),
    )
    metric_provenance["l0_complaint_cnt"] = _provenance(
        source_mode="SOURCE_MISSING",
        source_field=None,
        source_fields=["lesson_facts.complaint_level_rank"],
        batch_id=batch_id,
        note=(
            "The teacher-wide workbook has no complaint-level rows. L0 count must "
            "be derived from lesson_facts after exact level-3 complaint matching."
        ),
    )
    if policy_version in CAPACITY_MILESTONE_POLICY_VERSIONS:
        metric_provenance["peak_slot_cnt"] = _provenance(
            source_mode="SOURCE_MISSING" if peak_slot_value_missing else "REAL",
            source_field="peak_slot_cnt",
            batch_id=batch_id,
            note=(
                "Source peak_slot_cnt is empty; it is normalized to 0 for scoring and "
                "must not be presented as observed zero."
                if peak_slot_value_missing
                else "Direct teacher-wide Peak-slot count from the validated 61-column snapshot."
            ),
        )
        metric_provenance["capacity_score"] = _provenance(
            source_mode="DERIVED_REAL",
            source_field="peak_slot_cnt",
            source_fields=["peak_slot_cnt"],
            batch_id=batch_id,
            note=(
                "CAPACITY_PEAK_SLOT_40 awards 10 points at the first observed "
                "peak_slot_cnt >= 40 and then remains locked even if a later "
                "snapshot or correction falls below 40."
            ),
        )
    else:
        metric_provenance["capacity_score"] = _provenance(
            source_mode="MOCK",
            source_field=None,
            batch_id=batch_id,
            note="Historical v2/v3 fixed 10-point capacity assumption.",
        )
    metric_provenance["new_teacher_task_score"] = _provenance(
        source_mode="SOURCE_MISSING",
        source_field=None,
        batch_id=batch_id,
        note=(
            "No completed G01-G10 assignment evidence is connected for this teacher; "
            "the mandatory-growth score is zero."
        ),
    )
    metric_provenance["mandatory_task_completed_count"] = _provenance(
        source_mode="SOURCE_MISSING",
        source_field=None,
        source_fields=["task_assignments.status"],
        batch_id=batch_id,
        note=(
            "The teacher-wide workbook has no shared mandatory-task status. "
            "Completion count must be read from task_assignments."
        ),
    )
    metric_provenance["mandatory_task_assignment_count"] = deepcopy(
        metric_provenance["mandatory_task_completed_count"]
    )
    metric_provenance["mandatory_task_expected_count"] = _provenance(
        source_mode="SYSTEM_CONFIG",
        source_field="task_templates",
        batch_id=batch_id,
        note="The current mandatory catalog contains G01-G10.",
    )
    metric_provenance["class_quality_no_issue_rate"] = _provenance(
        source_mode="SOURCE_MISSING",
        source_field=None,
        batch_id=batch_id,
        note=(
            "Historical compatibility field. The current classroom-quality "
            "formula uses perfect_cnt and does not read this rate."
        ),
    )

    capacity_score = min(
        metric_inputs["capacity_score"], scoring_items["capacity"]["maximum_points"]
    )
    new_teacher_task_score = min(
        metric_inputs["new_teacher_task_score"],
        scoring_items["new_teacher_tasks"]["maximum_points"],
    )
    reliability_score = (
        on_time_completed * scoring_items["reliability_on_time"]["points_per_unit"]
        + peak_completed * scoring_items["reliability_peak"]["points_per_unit"]
    )
    feedback_score = (
        praise * scoring_items["feedback_praise"]["points_per_unit"]
        + favorite * scoring_items["feedback_favorite"]["points_per_unit"]
        + rebook * scoring_items["feedback_rebook_15d"]["points_per_unit"]
    )
    classroom_quality_rule = scoring_items["classroom_quality"]
    if classroom_quality_rule.get("metric") == "perfect_cnt":
        class_quality_score = perfect * classroom_quality_rule["points_per_unit"]
    else:
        class_quality_score = (
            total_completed
            * classroom_quality_rule["points_per_unit"]
            * metric_inputs["class_quality_no_issue_rate"]
        )
    raw_total_score = (
        reliability_score
        + feedback_score
        + class_quality_score
        + capacity_score
        + new_teacher_task_score
    )
    return {
        "metric_inputs": metric_inputs,
        "metric_provenance": metric_provenance,
        "reliability_score": round(float(reliability_score), 2),
        "user_feedback_score": round(float(feedback_score), 2),
        "class_quality_score": round(float(class_quality_score), 2),
        "raw_total_score": round(float(raw_total_score), 2),
        "public_total_score": _public_score(raw_total_score, effective_policy),
    }


def _dimensions(metrics: dict[str, Any]) -> list[dict[str, Any]]:
    inputs = metrics["metric_inputs"]
    return [
        {
            "code": "RELIABILITY",
            "label": "可靠性",
            "score": metrics["reliability_score"],
            "minimum": 0,
            "weight": 0,
            "data_mode": "DERIVED_REAL",
            "score_rule_version": SCORE_RULE_VERSION,
            "source_fields": [
                "total_completed_cnt",
                "late_cnt",
                "early_cnt",
                "peak_completed_cnt",
            ],
        },
        {
            "code": "USER_FEEDBACK",
            "label": "用户反馈",
            "score": metrics["user_feedback_score"],
            "minimum": 0,
            "weight": 0,
            "data_mode": "DERIVED_REAL",
            "score_rule_version": SCORE_RULE_VERSION,
            "source_fields": [
                "feedback_praise_cnt",
                "feedback_favorite_cnt",
                "completed_again_student_15d_cnt",
            ],
        },
        {
            "code": "CLASS_QUALITY",
            "label": "课堂质量",
            "score": metrics["class_quality_score"],
            "minimum": 0,
            "weight": 0,
            "data_mode": metrics["metric_provenance"]["perfect_cnt"]["source_mode"],
            "score_rule_version": SCORE_RULE_VERSION,
            "formula": "perfect_cnt * 1.6",
            "source_fields": ["perfect_cnt"],
        },
        {
            "code": "CAPACITY",
            "label": "供给达标（Peak slots）",
            "score": inputs["capacity_score"],
            "minimum": 0,
            "weight": 0,
            "data_mode": "DERIVED_REAL",
            "score_rule_version": SCORE_RULE_VERSION,
            "components": [
                {
                    "code": inputs["capacity_milestone_id"],
                    "metric": "peak_slot_cnt",
                    "value": inputs["peak_slot_cnt"],
                    "operator": "GTE",
                    "threshold": 40,
                    "score": inputs["capacity_score"],
                    "maximum_points": 10,
                    "milestone_achieved": inputs["capacity_milestone_achieved"],
                    "currently_meets_threshold": inputs[
                        "capacity_milestone_currently_meets_threshold"
                    ],
                    "settlement_mode": inputs["capacity_milestone_settlement_mode"],
                    "source_mode": "DERIVED_REAL",
                }
            ],
        },
        {
            "code": "NEW_TEACHER_TASK",
            "label": "成长任务（必修）",
            "score": inputs["new_teacher_task_score"],
            "minimum": 0,
            "weight": 0,
            "data_mode": "SOURCE_MISSING",
            "score_rule_version": SCORE_RULE_VERSION,
        },
    ]


def _graduation_criteria_met(metrics: dict[str, Any]) -> bool:
    inputs = metrics["metric_inputs"]
    thresholds = SCORE_POLICY_SNAPSHOT["thresholds"]
    gates = SCORE_POLICY_SNAPSHOT["hard_gates"]["graduation"]
    if SCORE_POLICY_SNAPSHOT["policy_version"] in {"v5", "v6"}:
        provenance = metrics["metric_provenance"]
        return bool(
            metrics["raw_total_score"] >= thresholds["graduation_raw_score"]
            and inputs["mandatory_task_completed_count"]
            == gates["required_mandatory_task_count"]
            and provenance["mandatory_task_completed_count"]["source_mode"]
            == "SYSTEM_TASK_STATUS"
            and inputs["l0_complaint_cnt"] <= gates["maximum_l0_complaint_count"]
            and provenance["l0_complaint_cnt"]["source_mode"]
            in {"REAL", "DERIVED_REAL"}
        )
    return bool(
        metrics["raw_total_score"] >= thresholds["graduation_raw_score"]
        # 2026-07-20 v3: the 30-point mandatory growth block must stand on
        # its own. Optional supply points cannot compensate for an incomplete
        # mandatory requirement.
        and inputs["new_teacher_task_score"] >= gates["minimum_base_score"]
        and inputs["total_completed_cnt"] >= gates["minimum_completed_lessons"]
        and metrics["user_feedback_score"]
        > gates["minimum_user_feedback_score_exclusive"]
        and metrics["reliability_score"]
        > gates["minimum_reliability_score_exclusive"]
        and (gates["allow_severe_redline"] or not inputs["severe_redline_event"])
    )


def _teacher_payload(
    row: _ValidatedRow,
    *,
    batch_id: str,
    snapshot_label: str,
    imported_at: datetime,
    existing_payload: dict[str, Any] | None,
    prior_capacity_milestone_achieved: bool = False,
) -> tuple[dict[str, Any], dict[str, Any]]:
    raw = row.raw_payload
    existing_metric_inputs = (
        (existing_payload or {}).get("metric_inputs")
        if isinstance((existing_payload or {}).get("metric_inputs"), dict)
        else {}
    )
    metrics = _metric_projection(
        raw,
        batch_id,
        row_number=row.row_number,
        prior_capacity_milestone_achieved=bool(
            prior_capacity_milestone_achieved
            or existing_metric_inputs.get("capacity_milestone_achieved", False)
        ),
    )
    inputs = metrics["metric_inputs"]
    name = str(raw.get("real_name") or row.teacher_id).strip()
    job_days = _as_nonnegative_int(raw, "job_days", row_number=row.row_number)
    onboard_date = _as_optional_date(raw, "onboard_date", row_number=row.row_number)
    onboard_end = _as_optional_date(raw, "onboard_30d_end_date", row_number=row.row_number)
    profile_projection = _teacher_profile_projection(
        raw,
        batch_id=batch_id,
        row_number=row.row_number,
    )
    first_booked_date = profile_projection["first_booked_date"]
    is_cpl_tesol = profile_projection["is_cpl_tesol"]
    is_self_introduce = profile_projection["is_self_introduce"]
    graduation_criteria_met = _graduation_criteria_met(metrics)
    thresholds = SCORE_POLICY_SNAPSHOT["thresholds"]
    gold_criteria_met = bool(
        graduation_criteria_met
        and metrics["raw_total_score"] >= thresholds["gold_raw_score"]
    )

    payload = deepcopy(existing_payload or {})
    payload.update(
        {
            "teacher_id": row.teacher_id,
            "name": name,
            "avatar": _avatar(name, row.teacher_id),
            "country": "Unknown",
            "timezone": None,
            "camp_day": min(max(job_days, 0), 30),
            "camp_window_state": "COMPLETED" if onboard_end is not None else "UNKNOWN",
            "employment_status": raw.get("status"),
            "bu": raw.get("bu"),
            "based_type": raw.get("based_type"),
            "teach_area_type": raw.get("teach_area_type"),
            "onboard_date": onboard_date.isoformat() if onboard_date else None,
            "onboard_30d_end_date": onboard_end.isoformat() if onboard_end else None,
            **deepcopy(profile_projection["payload"]),
            "lessons_completed": inputs["total_completed_cnt"],
            "data_mode": "MIXED",
            "source_batch_id": batch_id,
            "source_snapshot_label": snapshot_label,
            "score_rule_version": SCORE_RULE_VERSION,
            "score_policy_sha256": SCORE_POLICY_SHA256,
            "metric_inputs": deepcopy(inputs),
            "metric_provenance": deepcopy(metrics["metric_provenance"]),
            "capacity_milestone": {
                "milestone_id": inputs["capacity_milestone_id"],
                "metric": "peak_slot_cnt",
                "operator": "GTE",
                "threshold": 40,
                "current_value": inputs["peak_slot_cnt"],
                "currently_meets_threshold": inputs[
                    "capacity_milestone_currently_meets_threshold"
                ],
                "achieved": inputs["capacity_milestone_achieved"],
                "score": inputs["capacity_score"],
                "settlement_mode": inputs["capacity_milestone_settlement_mode"],
            },
            "profile_provenance": {
                "country": {
                    "source_mode": "SOURCE_MISSING",
                    "source_field": None,
                    "batch_id": batch_id,
                    "note": "The source workbook has no country column; Unknown is displayed without inference.",
                },
                "timezone": {
                    "source_mode": "SOURCE_MISSING",
                    "source_field": None,
                    "batch_id": batch_id,
                    "note": (
                        "The source workbook has no timezone column; payload remains null. "
                        "The non-null database column uses UTC only as an internal storage placeholder."
                    ),
                },
                **deepcopy(profile_projection["provenance"]),
            },
            "dimensions": _dimensions(metrics),
            "raw_total_score": metrics["raw_total_score"],
            "total_score": metrics["raw_total_score"],
            "external_display_score": metrics["public_total_score"],
            "graduation_threshold": thresholds["graduation_raw_score"],
            "gold_threshold": thresholds["gold_raw_score"],
            "graduation_external_score": thresholds["graduation_external_score"],
            "gold_external_score": thresholds["gold_external_score"],
            "score_policy_version": SCORE_POLICY_SNAPSHOT["policy_version"],
            "graduation_state": (
                "GRADUATED"
                if (
                    (existing_payload or {}).get("graduation_state") == "GRADUATED"
                    or graduation_criteria_met
                )
                else "IN_PROGRESS"
            ),
            "graduation_criteria_met": graduation_criteria_met,
            "gold_criteria_met": gold_criteria_met,
            "score_tier": (
                "GOLD" if gold_criteria_met else "GRADUATED" if graduation_criteria_met else "IN_PROGRESS"
            ),
            "signals": payload.get("signals", []),
            "lesson_facts": payload.get("lesson_facts", []),
            "lesson_dimension_scores": payload.get("lesson_dimension_scores", []),
            "score_entries": payload.get("score_entries", []),
            "risk_tags": payload.get("risk_tags", []),
            "updated_at": imported_at.isoformat(),
        }
    )
    snapshot_values = {
        "real_name": name,
        "employment_status": raw.get("status"),
        "bu": raw.get("bu"),
        "based_type": raw.get("based_type"),
        "teach_area_type": raw.get("teach_area_type"),
        "onboard_date": onboard_date,
        "onboard_30d_end_date": onboard_end,
        "first_booked_date": first_booked_date,
        "is_cpl_tesol": is_cpl_tesol,
        "is_self_introduce": is_self_introduce,
        "lessons_completed": inputs["total_completed_cnt"],
        "score_rule_version": SCORE_RULE_VERSION,
        "score_policy_snapshot": deepcopy(SCORE_POLICY_SNAPSHOT),
        "score_policy_sha256": SCORE_POLICY_SHA256,
        "total_completed_cnt": inputs["total_completed_cnt"],
        "peak_completed_cnt": inputs["peak_completed_cnt"],
        "peak_slot_cnt": inputs["peak_slot_cnt"],
        "perfect_cnt": inputs["perfect_cnt"],
        "on_time_completed_cnt": inputs["on_time_completed_cnt"],
        "feedback_praise_cnt": inputs["feedback_praise_cnt"],
        "feedback_favorite_cnt": inputs["feedback_favorite_cnt"],
        "completed_again_student_15d_cnt": inputs[
            "completed_again_student_15d_cnt"
        ],
        "late_cnt": inputs["late_cnt"],
        "early_cnt": inputs["early_cnt"],
        "real_absent_cnt": inputs["real_absent_cnt"],
        "severe_redline_event": inputs["severe_redline_event"],
        "capacity_score": inputs["capacity_score"],
        "new_teacher_task_score": inputs["new_teacher_task_score"],
        "class_quality_no_issue_rate": inputs["class_quality_no_issue_rate"],
        "reliability_score": metrics["reliability_score"],
        "user_feedback_score": metrics["user_feedback_score"],
        "class_quality_score": metrics["class_quality_score"],
        "raw_total_score": metrics["raw_total_score"],
        "public_total_score": metrics["public_total_score"],
        "metric_inputs": deepcopy(inputs),
        "metric_provenance": deepcopy(metrics["metric_provenance"]),
    }
    return payload, snapshot_values


def _validate_header(actual: Iterable[Any]) -> tuple[str, ...]:
    normalized = tuple("" if item is None else str(item).strip() for item in actual)
    if normalized == EXPECTED_HEADERS:
        return normalized
    missing = [item for item in EXPECTED_HEADERS if item not in normalized]
    unexpected = [item for item in normalized if item and item not in EXPECTED_HEADERS]
    first_mismatch = next(
        (
            index
            for index, (expected, received) in enumerate(
                zip(EXPECTED_HEADERS, normalized), start=1
            )
            if expected != received
        ),
        None,
    )
    raise ImportValidationError(
        "teacher workbook header must match the frozen 61-column contract exactly; "
        f"actual_count={len(normalized)}, missing={missing}, unexpected={unexpected}, "
        f"first_mismatch_column={first_mismatch}"
    )


def _read_and_validate(path: Path, sheet_name: str) -> tuple[list[_ValidatedRow], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ImportValidationError(
                f"required sheet {sheet_name!r} not found; available={workbook.sheetnames}"
            )
        worksheet = workbook[sheet_name]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            raise ImportValidationError("teacher workbook is empty")
        headers = _validate_header(header_row)
        seen: dict[str, int] = {}
        rows: list[_ValidatedRow] = []
        perfect_vs_on_time_difference_count = 0
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if all(value in (None, "") for value in values):
                continue
            if len(values) != len(EXPECTED_HEADERS):
                raise ImportValidationError(
                    f"row {row_number}: expected 61 cells, got {len(values)}"
                )
            raw = {header: _json_value(value) for header, value in zip(headers, values)}
            teacher_id = _teacher_id(raw["tchr_id"], row_number=row_number)
            if teacher_id in seen:
                raise ImportValidationError(
                    f"duplicate tchr_id {teacher_id!r} at rows {seen[teacher_id]} and {row_number}"
                )
            seen[teacher_id] = row_number
            perfect = _as_nonnegative_int(raw, "perfect_cnt", row_number=row_number)
            total = _as_nonnegative_int(raw, "total_completed_cnt", row_number=row_number)
            late = _as_nonnegative_int(raw, "late_cnt", row_number=row_number)
            early = _as_nonnegative_int(raw, "early_cnt", row_number=row_number)
            if perfect != max(total - late - early, 0):
                perfect_vs_on_time_difference_count += 1
            rows.append(
                _ValidatedRow(
                    row_number=row_number,
                    teacher_id=teacher_id,
                    raw_payload=raw,
                )
            )
        if not rows:
            raise ImportValidationError("teacher workbook has no data rows")
        return rows, perfect_vs_on_time_difference_count
    finally:
        workbook.close()


def _repair_existing_batch_policy_metadata(
    session: Session,
    batch: DataImportBatchRecord,
    rows: list[_ValidatedRow],
) -> None:
    """Repair policy lineage without rewriting the imported business facts.

    Content-addressed re-imports are normally no-ops.  They still need to be
    able to complete metadata introduced by a later schema migration, because
    otherwise an old batch can remain permanently unreadable by newer audit
    tooling.  The raw row and row-number checks below keep this repair path
    from becoming a silent data-replacement path.
    """

    stored_header = tuple(str(item) for item in (batch.header or []))
    if stored_header != EXPECTED_HEADERS:
        raise ImportValidationError(
            "stored batch header conflicts with the frozen 61-column contract"
        )

    snapshots = session.scalars(
        select(TeacherMetricSnapshotRecord).where(
            TeacherMetricSnapshotRecord.batch_id == batch.batch_id
        )
    ).all()
    expected_rows = {row.teacher_id: row for row in rows}
    snapshots_by_teacher = {snapshot.teacher_id: snapshot for snapshot in snapshots}
    if len(snapshots) != len(rows) or set(snapshots_by_teacher) != set(expected_rows):
        raise ImportValidationError(
            "stored batch snapshots are incomplete or contain unexpected teacher IDs; "
            "refusing to report an idempotent success"
        )

    now = _now()
    repaired_lineages: set[tuple[str, str]] = set()
    for teacher_id, snapshot in snapshots_by_teacher.items():
        source_row = expected_rows[teacher_id]
        if (
            snapshot.source_row_number != source_row.row_number
            or snapshot.raw_payload != source_row.raw_payload
        ):
            raise ImportValidationError(
                f"stored snapshot for teacher {teacher_id!r} conflicts with source content"
            )
        stored_policy = snapshot.score_policy_snapshot
        stored_policy_is_valid = False
        if isinstance(stored_policy, dict):
            try:
                ScoreGraduationConfig.model_validate(stored_policy)
                stored_policy_is_valid = (
                    score_policy_sha256(stored_policy) == snapshot.score_policy_sha256
                )
            except (TypeError, ValueError):
                stored_policy_is_valid = False
        effective_policy = stored_policy if stored_policy_is_valid else SCORE_POLICY_SNAPSHOT
        existing_milestone_entry = _find_capacity_milestone_entry(session, teacher_id)
        expected_metrics = _metric_projection(
            source_row.raw_payload,
            batch.batch_id,
            row_number=source_row.row_number,
            score_policy=effective_policy,
            prior_capacity_milestone_achieved=bool(
                existing_milestone_entry is not None
                or (snapshot.metric_inputs or {}).get(
                    "capacity_milestone_achieved", False
                )
            ),
        )
        expected_profile = _teacher_profile_projection(
            source_row.raw_payload,
            batch_id=batch.batch_id,
            row_number=source_row.row_number,
        )
        if (
            snapshot.metric_inputs != expected_metrics["metric_inputs"]
            or snapshot.metric_provenance != expected_metrics["metric_provenance"]
            or snapshot.reliability_score != expected_metrics["reliability_score"]
            or snapshot.user_feedback_score != expected_metrics["user_feedback_score"]
            or snapshot.class_quality_score != expected_metrics["class_quality_score"]
            or snapshot.raw_total_score != expected_metrics["raw_total_score"]
            or snapshot.public_total_score != expected_metrics["public_total_score"]
        ):
            raise ImportValidationError(
                f"stored scoring facts for teacher {teacher_id!r} cannot be reconstructed "
                "from the identical source content; refusing to relabel the policy"
            )
        if not stored_policy_is_valid:
            snapshot.score_rule_version = SCORE_RULE_VERSION
            snapshot.score_policy_snapshot = deepcopy(SCORE_POLICY_SNAPSHOT)
            snapshot.score_policy_sha256 = SCORE_POLICY_SHA256
            snapshot.updated_at = now
        if (
            snapshot.first_booked_date != expected_profile["first_booked_date"]
            or snapshot.is_cpl_tesol != expected_profile["is_cpl_tesol"]
            or snapshot.is_self_introduce != expected_profile["is_self_introduce"]
        ):
            snapshot.first_booked_date = expected_profile["first_booked_date"]
            snapshot.is_cpl_tesol = expected_profile["is_cpl_tesol"]
            snapshot.is_self_introduce = expected_profile["is_self_introduce"]
            snapshot.updated_at = now
        repaired_lineages.add((snapshot.score_rule_version, snapshot.score_policy_sha256))

        teacher = session.get(TeacherRecord, teacher_id)
        milestone_entry_payload: dict[str, Any] | None = None
        if (
            teacher is not None
            and bool(
                expected_metrics["metric_inputs"].get(
                    "capacity_milestone_achieved", False
                )
            )
        ):
            _, milestone_entry_payload, _ = _ensure_capacity_milestone_entry(
                session,
                teacher=teacher,
                batch_id=batch.batch_id,
                snapshot_id=snapshot.snapshot_id,
                peak_slot_cnt=int(
                    expected_metrics["metric_inputs"].get("peak_slot_cnt", 0)
                ),
                occurred_at=snapshot.created_at or batch.imported_at or now,
            )
        if teacher is None or teacher.source_batch_id != batch.batch_id:
            # A historical batch must not overwrite a teacher's newer current
            # projection.  The immutable snapshot above remains fully repaired.
            continue
        teacher_payload = deepcopy(teacher.payload or {})
        if milestone_entry_payload is not None:
            _append_score_entry_payload(teacher_payload, milestone_entry_payload)
        expected_profile_payload = expected_profile["payload"]
        profile_payload_changed = any(
            teacher_payload.get(field) != value
            for field, value in expected_profile_payload.items()
        )
        teacher_payload.update(deepcopy(expected_profile_payload))
        profile_provenance = deepcopy(teacher_payload.get("profile_provenance") or {})
        expected_profile_provenance = expected_profile["provenance"]
        profile_provenance_changed = any(
            profile_provenance.get(field) != value
            for field, value in expected_profile_provenance.items()
        )
        profile_provenance.update(deepcopy(expected_profile_provenance))
        teacher_payload["profile_provenance"] = profile_provenance
        graduation_criteria_met = _graduation_criteria_met(
            {
                "metric_inputs": snapshot.metric_inputs,
                "metric_provenance": snapshot.metric_provenance,
                "raw_total_score": snapshot.raw_total_score,
                "user_feedback_score": snapshot.user_feedback_score,
                "reliability_score": snapshot.reliability_score,
            }
        )
        repaired_graduation_state = (
            "GRADUATED"
            if teacher.graduation_state == "GRADUATED" or graduation_criteria_met
            else teacher.graduation_state
        )
        if (
            teacher_payload.get("score_rule_version") != snapshot.score_rule_version
            or teacher_payload.get("score_policy_sha256") != snapshot.score_policy_sha256
            or teacher_payload.get("graduation_state") != repaired_graduation_state
            or teacher.graduation_state != repaired_graduation_state
            or profile_payload_changed
            or profile_provenance_changed
        ):
            teacher_payload["score_rule_version"] = snapshot.score_rule_version
            teacher_payload["score_policy_sha256"] = snapshot.score_policy_sha256
            teacher_payload["graduation_state"] = repaired_graduation_state
            teacher.payload = teacher_payload
            teacher.graduation_state = repaired_graduation_state
            teacher.updated_at = now

    batch_payload = deepcopy(batch.payload or {})
    validation = deepcopy(batch_payload.get("validation") or {})
    if len(repaired_lineages) != 1:
        raise ImportValidationError(
            "stored batch contains multiple score-policy lineages; refusing an idempotent repair"
        )
    batch_rule_version, batch_policy_sha256 = next(iter(repaired_lineages))
    validation.update(
        {
            "score_rule_version": batch_rule_version,
            "score_policy_sha256": batch_policy_sha256,
        }
    )
    if batch_rule_version == SCORE_RULE_VERSION:
        validation.update(
            capacity_milestone_id=CAPACITY_MILESTONE_ID,
            capacity_milestone_settlement_mode=(
                CAPACITY_MILESTONE_SETTLEMENT_MODE
            ),
        )
    batch_payload["validation"] = validation
    if batch.payload != batch_payload:
        batch.payload = batch_payload
        batch.updated_at = now


def _replace_score_dimension(
    dimensions: Any,
    replacement: dict[str, Any],
) -> list[dict[str, Any]]:
    """Replace one score dimension and preserve the other ledgers."""

    existing = dimensions if isinstance(dimensions, list) else []
    replacement_code = replacement["code"]
    result: list[dict[str, Any]] = []
    replaced = False
    for dimension in existing:
        if (
            isinstance(dimension, dict)
            and dimension.get("code") == replacement_code
        ):
            if not replaced:
                result.append(deepcopy(replacement))
                replaced = True
            continue
        if isinstance(dimension, dict):
            result.append(deepcopy(dimension))
    if not replaced:
        result.append(deepcopy(replacement))
    return result


def recalculate_current_class_quality_scores(
    *,
    bind: Engine | None = None,
    dry_run: bool = True,
) -> ClassQualityRecalculationResult:
    """Rebuild the current quality projection from the immutable perfect count.

    Publishing a score policy does not mutate previously imported facts.  This
    explicit operation upgrades only each teacher's current snapshot and its
    current classroom-quality projections; historical non-current snapshots
    remain unchanged for audit.
    """

    selected_engine = bind or default_engine
    teachers_scanned = 0
    current_snapshots_found = 0
    snapshots_updated = 0
    teacher_payloads_updated = 0
    score_accounts_created = 0
    score_accounts_updated = 0
    teachers_without_current_snapshot = 0
    now = _now()

    with session_scope(selected_engine) as session:
        published_config = session.scalar(
            select(ConfigVersionRecord)
            .where(
                ConfigVersionRecord.config_key
                == ConfigKey.SCORE_GRADUATION.value,
                ConfigVersionRecord.status == ConfigStatus.PUBLISHED.value,
            )
            .order_by(ConfigVersionRecord.version_number.desc())
        )
        score_policy = (
            ScoreGraduationConfig.model_validate(published_config.payload).model_dump(
                mode="json"
            )
            if published_config is not None
            else deepcopy(SCORE_POLICY_SNAPSHOT)
        )
        classroom_quality_rule = score_policy["scoring_items"][
            "classroom_quality"
        ]
        if classroom_quality_rule.get("metric") != "perfect_cnt":
            raise RuntimeError(
                "published score policy must use perfect_cnt for classroom quality"
            )
        points_per_unit = float(classroom_quality_rule["points_per_unit"])
        if not math.isclose(points_per_unit, 1.6, abs_tol=1e-9):
            raise RuntimeError(
                "published classroom-quality policy must award 1.6 points per perfect lesson"
            )
        score_policy_digest = score_policy_sha256(score_policy)
        score_rule_version = (
            f"new_teacher_30d_20260723_{score_policy['policy_version']}"
        )

        teachers = session.scalars(
            select(TeacherRecord).order_by(TeacherRecord.teacher_id)
        ).all()
        teachers_scanned = len(teachers)
        current_teacher_ids = {
            teacher.teacher_id for teacher in teachers if teacher.source_batch_id
        }
        current_batch_ids = {
            str(teacher.source_batch_id)
            for teacher in teachers
            if teacher.source_batch_id
        }
        snapshots_by_current_key = {
            (snapshot.teacher_id, snapshot.batch_id): snapshot
            for snapshot in (
                session.scalars(
                    select(TeacherMetricSnapshotRecord).where(
                        TeacherMetricSnapshotRecord.teacher_id.in_(
                            current_teacher_ids
                        ),
                        TeacherMetricSnapshotRecord.batch_id.in_(
                            current_batch_ids
                        ),
                    )
                ).all()
                if current_teacher_ids and current_batch_ids
                else []
            )
        }
        quality_accounts_by_teacher = {
            account.teacher_id: account
            for account in (
                session.scalars(
                    select(ScoreAccountRecord).where(
                        ScoreAccountRecord.teacher_id.in_(current_teacher_ids),
                        ScoreAccountRecord.dimension == "CLASS_QUALITY",
                    )
                ).all()
                if current_teacher_ids
                else []
            )
        }
        task_accounts_by_teacher = {
            account.teacher_id: account
            for account in (
                session.scalars(
                    select(ScoreAccountRecord).where(
                        ScoreAccountRecord.teacher_id.in_(current_teacher_ids),
                        ScoreAccountRecord.dimension == "NEW_TEACHER_TASK",
                    )
                ).all()
                if current_teacher_ids
                else []
            )
        }
        task_counts_by_teacher = {
            str(teacher_id): {
                "assignment_count": int(assignment_count),
                "completed_count": int(completed_count or 0),
            }
            for teacher_id, assignment_count, completed_count in (
                session.execute(
                    select(
                        TaskAssignmentRecord.teacher_id,
                        func.count(TaskAssignmentRecord.assignment_id),
                        func.sum(
                            case(
                                (TaskAssignmentRecord.status == "COMPLETED", 1),
                                else_=0,
                            )
                        ),
                    )
                    .where(
                        TaskAssignmentRecord.teacher_id.in_(current_teacher_ids),
                        TaskAssignmentRecord.task_kind == "FIXED_GROWTH",
                        TaskAssignmentRecord.creator_system == "TRIGGER_CENTER",
                        TaskAssignmentRecord.source_mode == "REAL",
                        TaskAssignmentRecord.task_code.in_(
                            tuple(f"G{number:02d}" for number in range(1, 11))
                        ),
                    )
                    .group_by(TaskAssignmentRecord.teacher_id)
                ).all()
                if current_teacher_ids
                else []
            )
        }
        for teacher in teachers:
            if not teacher.source_batch_id:
                teachers_without_current_snapshot += 1
                continue
            snapshot = snapshots_by_current_key.get(
                (teacher.teacher_id, str(teacher.source_batch_id))
            )
            if snapshot is None:
                teachers_without_current_snapshot += 1
                continue
            current_snapshots_found += 1

            previous_inputs = (
                snapshot.metric_inputs
                if isinstance(snapshot.metric_inputs, dict)
                else {}
            )
            metrics = _metric_projection(
                snapshot.raw_payload,
                snapshot.batch_id,
                row_number=snapshot.source_row_number,
                score_policy=score_policy,
                prior_capacity_milestone_achieved=bool(
                    previous_inputs.get("capacity_milestone_achieved", False)
                ),
            )
            inputs = metrics["metric_inputs"]
            task_counts = task_counts_by_teacher.get(
                teacher.teacher_id,
                {"assignment_count": 0, "completed_count": 0},
            )
            task_account = task_accounts_by_teacher.get(teacher.teacher_id)
            task_account_payload = (
                task_account.payload
                if task_account is not None
                and isinstance(task_account.payload, dict)
                else {}
            )
            task_source_mode = str(
                task_account_payload.get("source_mode") or ""
            )
            if (
                task_counts["completed_count"] > 0
                and task_source_mode != "SYSTEM_TASK_STATUS"
            ):
                raise RuntimeError(
                    "completed mandatory tasks must be settled before score recalculation"
                )
            task_score = (
                max(float(task_account.current_score), 0)
                if task_account is not None
                and task_source_mode == "SYSTEM_TASK_STATUS"
                else 0.0
            )
            previous_task_score = float(
                inputs.get("new_teacher_task_score", 0)
            )
            inputs.update(
                {
                    "new_teacher_task_score": round(task_score, 2),
                    "mandatory_task_assignment_count": task_counts[
                        "assignment_count"
                    ],
                    "mandatory_task_completed_count": task_counts[
                        "completed_count"
                    ],
                    "mandatory_task_expected_count": 10,
                }
            )
            task_provenance = {
                "source_mode": (
                    "SYSTEM_TASK_STATUS"
                    if task_counts["assignment_count"] == 10
                    else "TASK_BASELINE_INCOMPLETE"
                ),
                "source_field": "task_assignments.status",
                "source_fields": [
                    "task_assignments.status",
                    "task_assignments.template_version_id",
                    "task_templates.payload.score_value",
                ],
                "batch_id": snapshot.batch_id,
                "note": (
                    "Mandatory-growth points are the configured values of "
                    "current COMPLETED G01-G10 assignments."
                ),
            }
            metrics["metric_provenance"].update(
                {
                    "new_teacher_task_score": deepcopy(task_provenance),
                    "mandatory_task_assignment_count": deepcopy(
                        task_provenance
                    ),
                    "mandatory_task_completed_count": deepcopy(
                        task_provenance
                    ),
                }
            )
            metrics["raw_total_score"] = round(
                float(metrics["raw_total_score"])
                - previous_task_score
                + task_score,
                2,
            )
            metrics["public_total_score"] = _public_score(
                metrics["raw_total_score"],
                score_policy,
            )
            snapshot_values = {
                "score_rule_version": score_rule_version,
                "score_policy_snapshot": deepcopy(score_policy),
                "score_policy_sha256": score_policy_digest,
                "total_completed_cnt": inputs["total_completed_cnt"],
                "peak_completed_cnt": inputs["peak_completed_cnt"],
                "peak_slot_cnt": inputs["peak_slot_cnt"],
                "perfect_cnt": inputs["perfect_cnt"],
                "on_time_completed_cnt": inputs["on_time_completed_cnt"],
                "feedback_praise_cnt": inputs["feedback_praise_cnt"],
                "feedback_favorite_cnt": inputs["feedback_favorite_cnt"],
                "completed_again_student_15d_cnt": inputs[
                    "completed_again_student_15d_cnt"
                ],
                "late_cnt": inputs["late_cnt"],
                "early_cnt": inputs["early_cnt"],
                "real_absent_cnt": inputs["real_absent_cnt"],
                "severe_redline_event": inputs["severe_redline_event"],
                "capacity_score": inputs["capacity_score"],
                "new_teacher_task_score": inputs["new_teacher_task_score"],
                "class_quality_no_issue_rate": inputs[
                    "class_quality_no_issue_rate"
                ],
                "reliability_score": metrics["reliability_score"],
                "user_feedback_score": metrics["user_feedback_score"],
                "class_quality_score": metrics["class_quality_score"],
                "raw_total_score": metrics["raw_total_score"],
                "public_total_score": metrics["public_total_score"],
                "metric_inputs": deepcopy(inputs),
                "metric_provenance": deepcopy(metrics["metric_provenance"]),
            }
            snapshot_changed = any(
                getattr(snapshot, field) != value
                for field, value in snapshot_values.items()
            )
            if snapshot_changed:
                snapshots_updated += 1
                if not dry_run:
                    for field, value in snapshot_values.items():
                        setattr(snapshot, field, value)
                    snapshot.updated_at = now

            quality_dimension = {
                "code": "CLASS_QUALITY",
                "label": "课堂质量",
                "score": metrics["class_quality_score"],
                "minimum": 0,
                "weight": 0,
                "data_mode": metrics["metric_provenance"]["perfect_cnt"][
                    "source_mode"
                ],
                "score_rule_version": score_rule_version,
                "formula": f"perfect_cnt * {points_per_unit:g}",
                "source_fields": ["perfect_cnt"],
            }
            task_dimension = {
                "code": "NEW_TEACHER_TASK",
                "label": "成长任务（必修）",
                "score": round(task_score, 2),
                "minimum": 0,
                "weight": 0,
                "data_mode": task_provenance["source_mode"],
                "source_mode": task_provenance["source_mode"],
                "score_rule_version": (
                    task_account.score_rule_version
                    if task_account is not None
                    and task_source_mode == "SYSTEM_TASK_STATUS"
                    else score_rule_version
                ),
            }
            teacher_payload = deepcopy(teacher.payload or {})
            candidate_payload = deepcopy(teacher_payload)
            candidate_payload.update(
                {
                    "score_rule_version": score_rule_version,
                    "score_policy_version": score_policy["policy_version"],
                    "score_policy_sha256": score_policy_digest,
                    "metric_inputs": deepcopy(inputs),
                    "metric_provenance": deepcopy(metrics["metric_provenance"]),
                    "dimensions": _replace_score_dimension(
                        _replace_score_dimension(
                            candidate_payload.get("dimensions"),
                            quality_dimension,
                        ),
                        task_dimension,
                    ),
                    "new_teacher_task_score": round(task_score, 2),
                    "base_score": round(
                        float(inputs["capacity_score"]) + task_score,
                        2,
                    ),
                    "raw_total_score": metrics["raw_total_score"],
                    "total_score": metrics["raw_total_score"],
                    "external_display_score": metrics["public_total_score"],
                }
            )
            teacher_changed = (
                candidate_payload != teacher_payload
                or not math.isclose(
                    float(teacher.total_score),
                    float(metrics["raw_total_score"]),
                    abs_tol=1e-9,
                )
            )
            if teacher_changed:
                teacher_payloads_updated += 1
                if not dry_run:
                    candidate_payload["updated_at"] = now.isoformat()
                    teacher.payload = candidate_payload
                    teacher.total_score = float(metrics["raw_total_score"])
                    teacher.updated_at = now

            account = quality_accounts_by_teacher.get(teacher.teacher_id)
            if account is None:
                score_accounts_created += 1
                if not dry_run:
                    session.add(
                        ScoreAccountRecord(
                            account_id=f"{teacher.teacher_id}:CLASS_QUALITY",
                            teacher_id=teacher.teacher_id,
                            camp_enrollment_id=teacher.camp_enrollment_id,
                            dimension="CLASS_QUALITY",
                            current_score=float(metrics["class_quality_score"]),
                            minimum_score=0,
                            weight=0,
                            score_rule_version=score_rule_version,
                            version=1,
                            updated_at=now,
                            payload=deepcopy(quality_dimension),
                        )
                    )
            else:
                account_changed = (
                    not math.isclose(
                        float(account.current_score),
                        float(metrics["class_quality_score"]),
                        abs_tol=1e-9,
                    )
                    or account.camp_enrollment_id != teacher.camp_enrollment_id
                    or account.minimum_score != 0
                    or account.weight != 0
                    or account.score_rule_version != score_rule_version
                    or account.payload != quality_dimension
                )
                if account_changed:
                    score_accounts_updated += 1
                    if not dry_run:
                        account.current_score = float(
                            metrics["class_quality_score"]
                        )
                        account.camp_enrollment_id = teacher.camp_enrollment_id
                        account.minimum_score = 0
                        account.weight = 0
                        account.score_rule_version = score_rule_version
                        account.version = int(account.version or 0) + 1
                        account.updated_at = now
                        account.payload = deepcopy(quality_dimension)

        if dry_run:
            session.rollback()

    return ClassQualityRecalculationResult(
        policy_version=str(score_policy["policy_version"]),
        score_policy_sha256=score_policy_digest,
        points_per_unit=points_per_unit,
        teachers_scanned=teachers_scanned,
        current_snapshots_found=current_snapshots_found,
        snapshots_updated=snapshots_updated,
        teacher_payloads_updated=teacher_payloads_updated,
        score_accounts_created=score_accounts_created,
        score_accounts_updated=score_accounts_updated,
        teachers_without_current_snapshot=teachers_without_current_snapshot,
        dry_run=dry_run,
    )


def import_teacher_metrics(
    source_path: str | Path,
    *,
    bind: Engine | None = None,
    snapshot_label: str | None = None,
    sheet_name: str = SOURCE_SHEET,
    source_system: str = SOURCE_SYSTEM,
    expected_sha256: str | None = None,
    expected_row_count: int | None = None,
) -> TeacherMetricImportResult:
    """Validate and idempotently upsert the delivered teacher-level XLSX snapshot."""

    path = Path(source_path).expanduser().resolve()
    if not path.is_file():
        raise ImportValidationError(f"source workbook does not exist: {path}")
    if path.suffix.casefold() != ".xlsx":
        raise ImportValidationError(f"source must be an .xlsx workbook: {path}")
    normalized_snapshot_label = (snapshot_label or path.stem).strip()
    if not normalized_snapshot_label or len(normalized_snapshot_label) > 128:
        raise ImportValidationError("snapshot_label must contain 1-128 characters")
    normalized_source_system = source_system.strip()
    if not normalized_source_system:
        raise ImportValidationError("source_system is required")

    source_sha256 = _sha256(path)
    if expected_sha256 and source_sha256 != expected_sha256.strip().casefold():
        raise ImportValidationError(
            f"source SHA-256 mismatch: expected={expected_sha256}, actual={source_sha256}"
        )
    rows, difference_count = _read_and_validate(path, sheet_name)
    if expected_row_count is not None and len(rows) != expected_row_count:
        raise ImportValidationError(
            f"source row-count mismatch: expected={expected_row_count}, actual={len(rows)}"
        )

    selected_engine = bind or default_engine
    batch_id = f"TMB-{source_sha256[:32]}"
    idempotent_reimport = False
    with session_scope(selected_engine) as session:
        existing_batch = session.scalar(
            select(DataImportBatchRecord).where(
                DataImportBatchRecord.source_sha256 == source_sha256,
                DataImportBatchRecord.source_sheet == sheet_name,
            )
        )
        if existing_batch is not None:
            idempotent_reimport = True
            batch_id = existing_batch.batch_id
            if existing_batch.snapshot_label != normalized_snapshot_label:
                raise ImportValidationError(
                    "identical source content was already imported with snapshot_label="
                    f"{existing_batch.snapshot_label!r}; refusing silent relabel to "
                    f"{normalized_snapshot_label!r}"
                )
            if existing_batch.source_system != normalized_source_system:
                raise ImportValidationError(
                    "identical source content was already imported from source_system="
                    f"{existing_batch.source_system!r}; received {normalized_source_system!r}"
                )
            if existing_batch.row_count != len(rows) or existing_batch.column_count != len(
                EXPECTED_HEADERS
            ):
                raise ImportValidationError("stored batch metadata conflicts with validated source content")
            _repair_existing_batch_policy_metadata(session, existing_batch, rows)
            try:
                ensure_fixed_growth_assignments(
                    session,
                    (row.teacher_id for row in rows),
                    occurred_at=existing_batch.imported_at,
                )
            except FixedGrowthBaselineError as exc:
                raise ImportValidationError(str(exc)) from exc
            return TeacherMetricImportResult(
                batch_id=existing_batch.batch_id,
                source_uri=str(path),
                source_sha256=source_sha256,
                source_sheet=sheet_name,
                snapshot_label=normalized_snapshot_label,
                row_count=len(rows),
                column_count=len(EXPECTED_HEADERS),
                data_mode="MIXED",
                idempotent_reimport=True,
                perfect_vs_on_time_difference_count=difference_count,
            )
        else:
            imported_at = _now()
            session.add(
                DataImportBatchRecord(
                    batch_id=batch_id,
                    source_system=normalized_source_system,
                    source_filename=path.name,
                    source_uri=str(path),
                    source_sha256=source_sha256,
                    source_sheet=sheet_name,
                    snapshot_label=normalized_snapshot_label,
                    data_mode="MIXED",
                    column_count=len(EXPECTED_HEADERS),
                    row_count=len(rows),
                    header=list(EXPECTED_HEADERS),
                    status="COMPLETED",
                    imported_at=imported_at,
                    payload={
                        "validation": {
                            "header_contract": "EXACT_61_COLUMNS",
                            "teacher_id_unique": True,
                            "row_count": len(rows),
                            "source_sha256": source_sha256,
                            "perfect_vs_on_time_difference_count": difference_count,
                            "on_time_formula": (
                                "max(total_completed_cnt-late_cnt-early_cnt,0)"
                            ),
                            "score_rule_version": SCORE_RULE_VERSION,
                            "score_policy_sha256": SCORE_POLICY_SHA256,
                            "capacity_milestone_id": CAPACITY_MILESTONE_ID,
                            "capacity_milestone_settlement_mode": (
                                CAPACITY_MILESTONE_SETTLEMENT_MODE
                            ),
                        },
                        "data_mode": "MIXED",
                    },
                    created_at=imported_at,
                    updated_at=imported_at,
                )
            )
            session.flush()

        for row in rows:
            existing_teacher = session.get(TeacherRecord, row.teacher_id)
            if existing_teacher is not None and existing_teacher.data_mode == "MOCK":
                raise ImportValidationError(
                    f"source tchr_id {row.teacher_id!r} collides with a protected Mock demo teacher"
                )
            existing_milestone_entry = _find_capacity_milestone_entry(
                session, row.teacher_id
            )
            teacher_payload, snapshot = _teacher_payload(
                row,
                batch_id=batch_id,
                snapshot_label=normalized_snapshot_label,
                imported_at=imported_at,
                existing_payload=existing_teacher.payload if existing_teacher else None,
                prior_capacity_milestone_achieved=existing_milestone_entry is not None,
            )
            camp_enrollment_id = (
                existing_teacher.camp_enrollment_id
                if existing_teacher is not None
                else f"CAMP-{row.teacher_id}"
            )
            teacher_record = session.merge(
                TeacherRecord(
                    teacher_id=row.teacher_id,
                    camp_enrollment_id=camp_enrollment_id,
                    name=teacher_payload["name"],
                    country=teacher_payload["country"],
                    timezone=teacher_payload.get("timezone") or "UTC",
                    camp_day=teacher_payload["camp_day"],
                    graduation_state=teacher_payload["graduation_state"],
                    total_score=float(teacher_payload["total_score"]),
                    graduation_threshold=float(teacher_payload["graduation_threshold"]),
                    data_mode="MIXED",
                    source_batch_id=batch_id,
                    source_snapshot_label=normalized_snapshot_label,
                    payload=deepcopy(teacher_payload),
                    created_at=existing_teacher.created_at if existing_teacher else imported_at,
                    updated_at=imported_at,
                )
            )
            if bool(snapshot["metric_inputs"]["capacity_milestone_achieved"]):
                _, milestone_entry_payload, _ = _ensure_capacity_milestone_entry(
                    session,
                    teacher=teacher_record,
                    batch_id=batch_id,
                    snapshot_id=f"{batch_id}:{row.teacher_id}",
                    peak_slot_cnt=int(snapshot["peak_slot_cnt"]),
                    occurred_at=imported_at,
                )
                _append_score_entry_payload(teacher_payload, milestone_entry_payload)
                teacher_record.payload = deepcopy(teacher_payload)
            existing_snapshot = session.get(
                TeacherMetricSnapshotRecord, f"{batch_id}:{row.teacher_id}"
            )
            session.merge(
                TeacherMetricSnapshotRecord(
                    snapshot_id=f"{batch_id}:{row.teacher_id}",
                    batch_id=batch_id,
                    teacher_id=row.teacher_id,
                    snapshot_label=normalized_snapshot_label,
                    source_row_number=row.row_number,
                    data_mode="MIXED",
                    raw_payload=deepcopy(row.raw_payload),
                    created_at=existing_snapshot.created_at if existing_snapshot else imported_at,
                    updated_at=imported_at,
                    **snapshot,
                )
            )
            for dimension in teacher_payload["dimensions"]:
                session.merge(
                    ScoreAccountRecord(
                        account_id=f'{row.teacher_id}:{dimension["code"]}',
                        teacher_id=row.teacher_id,
                        camp_enrollment_id=camp_enrollment_id,
                        dimension=dimension["code"],
                        current_score=float(dimension["score"]),
                        minimum_score=float(dimension.get("minimum", 0)),
                        weight=float(dimension.get("weight", 0)),
                        score_rule_version=SCORE_RULE_VERSION,
                        version=1,
                        updated_at=imported_at,
                        payload=deepcopy(dimension),
                    )
                )
        try:
            ensure_fixed_growth_assignments(
                session,
                (row.teacher_id for row in rows),
                occurred_at=imported_at,
            )
        except FixedGrowthBaselineError as exc:
            raise ImportValidationError(str(exc)) from exc

    return TeacherMetricImportResult(
        batch_id=batch_id,
        source_uri=str(path),
        source_sha256=source_sha256,
        source_sheet=sheet_name,
        snapshot_label=normalized_snapshot_label,
        row_count=len(rows),
        column_count=len(EXPECTED_HEADERS),
        data_mode="MIXED",
        idempotent_reimport=idempotent_reimport,
        perfect_vs_on_time_difference_count=difference_count,
    )
