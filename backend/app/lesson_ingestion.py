from __future__ import annotations

import hashlib
import json
import math
import re
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import load_workbook
from sqlalchemy import Engine, select
from sqlalchemy.orm import Session

from .database import engine as default_engine
from .db_models import (
    ComplaintCategoryRuleRecord,
    DataImportBatchRecord,
    LessonFactRecord,
    NotificationRecord,
    OpsCaseRecord,
    PersonalizedTriggerMatchRecord,
    SourceRecord,
    TaskAssignmentRecord,
    TaskTemplateRecord,
    TeacherRecord,
)
from .personalized_rules import (
    ComplaintRule,
    TriggerDecision,
    evaluate_lesson,
    normalize_text,
)


LESSON_SOURCE_SYSTEM = "MANUAL_XLSX:NEW_TEACHER_30D_LESSONS"
COMPLAINT_SOURCE_SYSTEM = "MANUAL_XLSX:COMPLAINT_LEVEL_RULES"
COMPLAINT_SOURCE_SHEET = "客服&销售&学员端投诉分级"
COMPLAINT_SOURCE_REGION = "A1:F45"
EXPECTED_LESSON_ROW_COUNT = 37_317
TRIGGER_RULE_VERSION = "personalized_rules_20260724_v2"

# The first real lesson baseline is a strict source contract. A renamed,
# omitted, reordered or additional column must be reviewed before import.
EXPECTED_LESSON_HEADERS: tuple[str, ...] = (
    "课程id",
    "上课日期",
    "上课时间",
    "是否高峰",
    "老师id",
    "学员id",
    "课程状态",
    "缺席原因明细",
    "迟到",
    "早退",
    "差评分",
    "差评标签",
    "投诉一级分类",
    "投诉二级分类",
    "投诉三级分类",
    "是否拉黑",
    "收藏",
    "好评标签",
    "评价详情",
    "是否复约",
    "未开摄像头",
    "cpu占用过高",
    "网络延迟过高",
    "假早退",
)

EXPECTED_COMPLAINT_HEADERS: tuple[str, ...] = (
    "一级分类",
    "二级分类",
    "三级分类",
    "P级",
    "Course Title in the Learning Hub",
    "link",
)

PERSONALIZED_TEMPLATE_CODES = {
    "P-REL-MEMO",
    "P-REL-ATTENDANCE",
    "P-FB-NEGATIVE",
    "P-FB-COMPLAINT",
    "P-FB-BLACKLIST",
}


class LessonImportValidationError(ValueError):
    """Raised before commit when a workbook or database preflight fails."""


@dataclass(frozen=True)
class LessonBaselineImportResult:
    lesson_batch_id: str
    complaint_batch_id: str
    lesson_source_sha256: str
    complaint_source_sha256: str
    lesson_sheet: str
    lesson_row_count: int
    complaint_source_row_count: int
    complaint_rule_count: int
    complaint_unmapped_source_rows: list[int]
    source_records_created: int
    lesson_facts_created: int
    task_assignments_created: int
    ops_cases_created: int
    notifications_created: int
    pending_data_matches_created: int
    trigger_matches_created: int
    blacklist_tasks_created: int
    negative_tag_pending_teachers: int
    unmatched_complaint_lessons: int
    trigger_counts: dict[str, int]
    dry_run: bool
    idempotent_reimport: bool

    def as_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class _ComplaintRow:
    row_number: int
    raw_payload: dict[str, Any]
    category_l1: str
    category_l2: str
    category_l3: str | None
    source_level: str
    normalized_level: str
    severity_rank: int
    default_route: str
    learning_title: str | None
    learning_url: str | None


@dataclass(frozen=True)
class _LessonRow:
    row_number: int
    raw_payload: dict[str, Any]
    lesson_id: str
    teacher_id: str
    student_id: str
    local_date: date
    local_time: time
    local_start_at: datetime
    lifecycle_status: str
    is_peak: bool | None
    is_late: bool | None
    is_early: bool | None
    is_false_early_leave: bool | None
    negative_score: float | None
    has_negative_tag: bool | None
    feedback_detail: str | None
    negative_tags: tuple[str, ...]
    absence_reason_detail: str | None
    complaint_l1: str | None
    complaint_l2: str | None
    complaint_l3: str | None
    is_blocked: bool | None
    is_favorited: bool | None
    has_positive_tag: bool | None
    is_rebooked: bool | None
    is_camera_off: bool | None
    is_cpu_usage_high: bool | None
    is_network_delay_high: bool | None


@dataclass(frozen=True)
class _OutputSpec:
    rule_code: str
    domain: str
    output_type: str
    title: str
    priority: str
    why: str
    evidence: dict[str, Any]
    teacher_id: str
    lesson_id: str | None
    source_record_id: str | None
    complaint_rule_id: str | None
    scope_key: str
    dedupe_key: str
    output_dedupe_key: str
    task_code: str | None = None


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _stable_id(prefix: str, value: str, *, length: int = 32) -> str:
    digest = hashlib.sha256(value.encode("utf-8")).hexdigest()[:length]
    return f"{prefix}-{digest}"


def _json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        raise LessonImportValidationError("source contains a non-finite number")
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _canonical_json(value: Any) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    )


def _identifier(value: Any, *, field: str, row_number: int, max_length: int) -> str:
    if value in (None, "") or isinstance(value, bool):
        raise LessonImportValidationError(f"row {row_number}: {field} is required")
    if isinstance(value, float):
        if not value.is_integer():
            raise LessonImportValidationError(
                f"row {row_number}: {field} must be an integer identifier"
            )
        value = int(value)
    normalized = str(value).strip()
    if re.fullmatch(r"[+-]?\d+\.0+", normalized):
        normalized = normalized.split(".", 1)[0]
    if not normalized:
        raise LessonImportValidationError(f"row {row_number}: {field} is required")
    if len(normalized) > max_length:
        raise LessonImportValidationError(
            f"row {row_number}: {field} exceeds {max_length} characters"
        )
    return normalized


def _optional_text(value: Any) -> str | None:
    normalized = normalize_text(value)
    return normalized or None


def _feedback_labels(value: Any) -> tuple[str, ...]:
    """Split the reviewed source's ASCII-comma label list, preserving order."""

    labels = [
        normalize_text(item)
        for item in str(value or "").split(",")
    ]
    return tuple(dict.fromkeys(item for item in labels if item))


def _required_text(value: Any, *, field: str, row_number: int) -> str:
    normalized = normalize_text(value)
    if not normalized:
        raise LessonImportValidationError(f"row {row_number}: {field} is required")
    return normalized


def _optional_bool(value: Any, *, field: str, row_number: int) -> bool | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool) and value in (0, 1):
        return bool(int(value))
    if isinstance(value, str) and value.strip() in {"0", "1"}:
        return value.strip() == "1"
    raise LessonImportValidationError(
        f"row {row_number}: {field} must be a controlled 0/1 value, got {value!r}"
    )


def _optional_float(value: Any, *, field: str, row_number: int) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        raise LessonImportValidationError(
            f"row {row_number}: {field} cannot be boolean"
        )
    try:
        converted = float(value)
    except (TypeError, ValueError) as exc:
        raise LessonImportValidationError(
            f"row {row_number}: {field} must be numeric, got {value!r}"
        ) from exc
    if not math.isfinite(converted):
        raise LessonImportValidationError(
            f"row {row_number}: {field} must be finite"
        )
    return converted


def _local_date(value: Any, *, row_number: int) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    normalized = normalize_text(value)
    try:
        return date.fromisoformat(normalized)
    except ValueError as exc:
        raise LessonImportValidationError(
            f"row {row_number}: 上课日期 must be an Excel/ISO date, got {value!r}"
        ) from exc


def _local_time_and_datetime(
    value: Any,
    *,
    lesson_date: date,
    row_number: int,
) -> tuple[time, datetime]:
    parsed_time: time
    if isinstance(value, datetime):
        if value.date() != lesson_date:
            raise LessonImportValidationError(
                f"row {row_number}: 上课时间 date conflicts with 上课日期"
            )
        parsed_time = value.time().replace(tzinfo=None)
    elif isinstance(value, time):
        parsed_time = value.replace(tzinfo=None)
    else:
        normalized = normalize_text(value)
        try:
            if "T" in normalized or " " in normalized:
                parsed = datetime.fromisoformat(normalized)
                if parsed.date() != lesson_date:
                    raise ValueError
                parsed_time = parsed.time().replace(tzinfo=None)
            else:
                parsed_time = time.fromisoformat(normalized).replace(tzinfo=None)
        except ValueError as exc:
            raise LessonImportValidationError(
                f"row {row_number}: 上课时间 must be an Excel/ISO local time, got {value!r}"
            ) from exc
    return parsed_time, datetime.combine(lesson_date, parsed_time)


def _read_lesson_workbook(
    source: Path,
    *,
    expected_row_count: int | None,
) -> tuple[str, str, list[_LessonRow]]:
    if not source.is_file():
        raise LessonImportValidationError(f"lesson workbook does not exist: {source}")
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        if len(workbook.sheetnames) != 1:
            raise LessonImportValidationError(
                "lesson workbook must contain exactly one reviewed source sheet"
            )
        sheet_name = workbook.sheetnames[0]
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = tuple(_json_value(value) for value in next(rows))
        except StopIteration as exc:
            raise LessonImportValidationError("lesson workbook is empty") from exc
        if headers != EXPECTED_LESSON_HEADERS:
            raise LessonImportValidationError(
                "lesson header contract mismatch: "
                f"expected {EXPECTED_LESSON_HEADERS!r}, got {headers!r}"
            )

        result: list[_LessonRow] = []
        seen_lesson_ids: set[str] = set()
        for row_number, values in enumerate(rows, start=2):
            if all(value in (None, "") for value in values):
                raise LessonImportValidationError(
                    f"row {row_number}: blank rows are not allowed inside the baseline"
                )
            if len(values) != len(EXPECTED_LESSON_HEADERS):
                raise LessonImportValidationError(
                    f"row {row_number}: expected {len(EXPECTED_LESSON_HEADERS)} cells, "
                    f"got {len(values)}"
                )
            raw = {
                field: _json_value(value)
                for field, value in zip(EXPECTED_LESSON_HEADERS, values)
            }
            source_values = dict(zip(EXPECTED_LESSON_HEADERS, values))
            has_negative_tag = _optional_bool(
                source_values["差评标签"],
                field="差评标签",
                row_number=row_number,
            )
            feedback_detail = _optional_text(source_values["评价详情"])
            lesson_id = _identifier(
                source_values["课程id"],
                field="课程id",
                row_number=row_number,
                max_length=128,
            )
            if lesson_id in seen_lesson_ids:
                raise LessonImportValidationError(
                    f"row {row_number}: duplicate 课程id {lesson_id}"
                )
            seen_lesson_ids.add(lesson_id)
            lesson_date = _local_date(source_values["上课日期"], row_number=row_number)
            lesson_time, local_start_at = _local_time_and_datetime(
                source_values["上课时间"],
                lesson_date=lesson_date,
                row_number=row_number,
            )
            result.append(
                _LessonRow(
                    row_number=row_number,
                    raw_payload=raw,
                    lesson_id=lesson_id,
                    teacher_id=_identifier(
                        source_values["老师id"],
                        field="老师id",
                        row_number=row_number,
                        max_length=64,
                    ),
                    student_id=_identifier(
                        source_values["学员id"],
                        field="学员id",
                        row_number=row_number,
                        max_length=128,
                    ),
                    local_date=lesson_date,
                    local_time=lesson_time,
                    local_start_at=local_start_at,
                    lifecycle_status=_required_text(
                        source_values["课程状态"],
                        field="课程状态",
                        row_number=row_number,
                    ),
                    is_peak=_optional_bool(
                        source_values["是否高峰"],
                        field="是否高峰",
                        row_number=row_number,
                    ),
                    is_late=_optional_bool(
                        source_values["迟到"], field="迟到", row_number=row_number
                    ),
                    is_early=_optional_bool(
                        source_values["早退"], field="早退", row_number=row_number
                    ),
                    is_false_early_leave=_optional_bool(
                        source_values["假早退"], field="假早退", row_number=row_number
                    ),
                    negative_score=_optional_float(
                        source_values["差评分"], field="差评分", row_number=row_number
                    ),
                    has_negative_tag=has_negative_tag,
                    feedback_detail=feedback_detail,
                    negative_tags=(
                        _feedback_labels(feedback_detail)
                        if has_negative_tag
                        else ()
                    ),
                    absence_reason_detail=_optional_text(source_values["缺席原因明细"]),
                    complaint_l1=_optional_text(source_values["投诉一级分类"]),
                    complaint_l2=_optional_text(source_values["投诉二级分类"]),
                    complaint_l3=_optional_text(source_values["投诉三级分类"]),
                    is_blocked=_optional_bool(
                        source_values["是否拉黑"], field="是否拉黑", row_number=row_number
                    ),
                    is_favorited=_optional_bool(
                        source_values["收藏"], field="收藏", row_number=row_number
                    ),
                    has_positive_tag=_optional_bool(
                        source_values["好评标签"], field="好评标签", row_number=row_number
                    ),
                    is_rebooked=_optional_bool(
                        source_values["是否复约"], field="是否复约", row_number=row_number
                    ),
                    is_camera_off=_optional_bool(
                        source_values["未开摄像头"], field="未开摄像头", row_number=row_number
                    ),
                    is_cpu_usage_high=_optional_bool(
                        source_values["cpu占用过高"],
                        field="cpu占用过高",
                        row_number=row_number,
                    ),
                    is_network_delay_high=_optional_bool(
                        source_values["网络延迟过高"],
                        field="网络延迟过高",
                        row_number=row_number,
                    ),
                )
            )
        if expected_row_count is not None and len(result) != expected_row_count:
            raise LessonImportValidationError(
                f"lesson row-count mismatch: expected {expected_row_count}, got {len(result)}"
            )
        return sheet_name, _sha256(source), result
    finally:
        workbook.close()


def _complaint_route(level2: str, severity_rank: int) -> str:
    if level2 == "出席问题":
        return "RELIABILITY"
    if level2 == "网络设备问题":
        return "CLASS_QUALITY"
    if severity_rank <= 1:
        return "OPS_CASE"
    return "USER_FEEDBACK"


def _read_complaint_workbook(
    source: Path,
) -> tuple[str, list[_ComplaintRow], list[int], list[dict[str, Any]]]:
    if not source.is_file():
        raise LessonImportValidationError(f"complaint workbook does not exist: {source}")
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        if COMPLAINT_SOURCE_SHEET not in workbook.sheetnames:
            raise LessonImportValidationError(
                f"complaint sheet is missing: {COMPLAINT_SOURCE_SHEET}"
            )
        worksheet = workbook[COMPLAINT_SOURCE_SHEET]
        region = list(
            worksheet.iter_rows(
                min_row=1,
                max_row=45,
                min_col=1,
                max_col=6,
                values_only=True,
            )
        )
        if len(region) != 45 or normalize_text(region[0][0]) != "现行版本":
            raise LessonImportValidationError(
                f"complaint source region {COMPLAINT_SOURCE_REGION} is not the reviewed current version"
            )
        headers = tuple(_json_value(value) for value in region[1])
        if headers != EXPECTED_COMPLAINT_HEADERS:
            raise LessonImportValidationError(
                "complaint header contract mismatch: "
                f"expected {EXPECTED_COMPLAINT_HEADERS!r}, got {headers!r}"
            )

        current_l1 = ""
        current_l2 = ""
        rules: list[_ComplaintRow] = []
        skipped_rows: list[int] = []
        raw_rows: list[dict[str, Any]] = []
        normalized_seen: dict[str, tuple[str, int]] = {}
        for row_number, values in enumerate(region[2:], start=3):
            raw = {
                field: _json_value(value)
                for field, value in zip(EXPECTED_COMPLAINT_HEADERS, values)
            }
            raw_rows.append(raw)
            current_l1 = normalize_text(values[0]) or current_l1
            current_l2 = normalize_text(values[1]) or current_l2
            if not current_l1 or not current_l2:
                raise LessonImportValidationError(
                    f"complaint row {row_number}: level-1/level-2 fill-down has no source value"
                )
            level_match = re.fullmatch(r"P([0-4])", normalize_text(values[3]).upper())
            if level_match is None:
                raise LessonImportValidationError(
                    f"complaint row {row_number}: P级 must be P0-P4, got {values[3]!r}"
                )
            source_level = f"P{level_match.group(1)}"
            rank = int(level_match.group(1))
            category_l3 = _optional_text(values[2])
            if category_l3 is None:
                skipped_rows.append(row_number)
                continue
            normalized_l3 = normalize_text(category_l3)
            previous = normalized_seen.get(normalized_l3)
            current = (source_level, row_number)
            if previous is not None:
                raise LessonImportValidationError(
                    "complaint third-level category is duplicated: "
                    f"{category_l3!r} at rows {previous[1]} and {row_number}"
                )
            normalized_seen[normalized_l3] = current
            rules.append(
                _ComplaintRow(
                    row_number=row_number,
                    raw_payload=raw,
                    category_l1=current_l1,
                    category_l2=current_l2,
                    category_l3=category_l3,
                    source_level=source_level,
                    normalized_level=f"L{rank}",
                    severity_rank=rank,
                    default_route=_complaint_route(current_l2, rank),
                    learning_title=_optional_text(values[4]),
                    learning_url=_optional_text(values[5]),
                )
            )
        if len(raw_rows) != 43:
            raise LessonImportValidationError("complaint source must contain rows 3-45")
        return _sha256(source), rules, skipped_rows, raw_rows
    finally:
        workbook.close()


def _source_record_id(batch_id: str, row_number: int) -> str:
    return f"SRC-{batch_id}-{row_number}"


def _complaint_rule_id(batch_id: str, row_number: int) -> str:
    return f"CR-{batch_id}-{row_number}"


def _chunks(values: Sequence[str], size: int = 900) -> Iterable[Sequence[str]]:
    for index in range(0, len(values), size):
        yield values[index : index + size]


def _existing_values(
    session: Session,
    column: Any,
    values: Iterable[str],
) -> set[str]:
    unique = sorted(set(values))
    found: set[str] = set()
    for chunk in _chunks(unique):
        found.update(session.scalars(select(column).where(column.in_(chunk))).all())
    return found


def _teacher_map(session: Session, teacher_ids: Iterable[str]) -> dict[str, TeacherRecord]:
    unique = sorted(set(teacher_ids))
    result: dict[str, TeacherRecord] = {}
    for chunk in _chunks(unique):
        for teacher in session.scalars(
            select(TeacherRecord).where(TeacherRecord.teacher_id.in_(chunk))
        ).all():
            result[teacher.teacher_id] = teacher
    missing = sorted(set(unique) - set(result))
    if missing:
        sample = missing[:10]
        raise LessonImportValidationError(
            "lesson preflight failed: teacher foreign keys are missing; "
            f"missing_count={len(missing)}, sample={sample!r}"
        )
    return result


def _template_map(session: Session) -> dict[str, TaskTemplateRecord]:
    templates = {
        item.template_id: item
        for item in session.scalars(
            select(TaskTemplateRecord).where(
                TaskTemplateRecord.template_id.in_(sorted(PERSONALIZED_TEMPLATE_CODES)),
                TaskTemplateRecord.status == "PUBLISHED",
            )
        ).all()
    }
    missing = sorted(PERSONALIZED_TEMPLATE_CODES - set(templates))
    if missing:
        raise LessonImportValidationError(
            f"personalized task templates must be seeded and published first: {missing!r}"
        )
    return templates


def _complaint_rule_context(
    rules: Sequence[_ComplaintRow],
    *,
    complaint_batch_id: str,
) -> tuple[dict[str, ComplaintRule], dict[str, str]]:
    evaluator_rules: dict[str, ComplaintRule] = {}
    rule_ids: dict[str, str] = {}
    for item in rules:
        assert item.category_l3 is not None
        normalized_l3 = normalize_text(item.category_l3)
        evaluator_rules[normalized_l3] = ComplaintRule(
            level2_name=item.category_l2,
            level3_name=item.category_l3,
            source_level_code=item.source_level,
            severity_rank=item.severity_rank,
            route_domain=item.default_route,
        )
        rule_ids[normalized_l3] = _complaint_rule_id(
            complaint_batch_id, item.row_number
        )
    return evaluator_rules, rule_ids


def _priority_min(values: Iterable[str]) -> str:
    return min(values, key=lambda item: int(item[1:]))


def _merge_lesson_decisions(
    row: _LessonRow,
    decisions: Sequence[TriggerDecision],
    *,
    source_record_id: str,
    complaint_rule_id: str | None,
) -> list[_OutputSpec]:
    grouped: dict[tuple[str, str], list[TriggerDecision]] = defaultdict(list)
    for decision in decisions:
        if decision.output_type == "TEACHER_TASK":
            key = (decision.output_type, decision.task_code or decision.rule_code)
        elif decision.output_type == "NOTIFICATION" and decision.domain == "CLASS_QUALITY":
            key = (decision.output_type, "CLASS_QUALITY")
        else:
            key = (decision.output_type, decision.rule_code)
        grouped[key].append(decision)

    result: list[_OutputSpec] = []
    for (output_type, key), members in grouped.items():
        task_code = members[0].task_code
        if output_type == "TEACHER_TASK" and task_code:
            canonical_rule = {
                "P-REL-MEMO": "TR-REL-LESSON-MEMO",
                "P-REL-ATTENDANCE": "TR-REL-ATTENDANCE",
                "P-FB-COMPLAINT": "TR-FB-GENERAL-COMPLAINT",
            }.get(task_code, members[0].rule_code)
        elif output_type == "NOTIFICATION" and key == "CLASS_QUALITY":
            canonical_rule = "TR-QUALITY-LESSON"
        else:
            canonical_rule = members[0].rule_code

        signals = [
            {
                "rule_code": item.rule_code,
                "why": item.why,
                "evidence": item.evidence,
            }
            for item in members
        ]
        evidence: dict[str, Any] = {
            "lesson_id": row.lesson_id,
            "source_row_number": row.row_number,
            "matched_rule_codes": list(dict.fromkeys(item.rule_code for item in members)),
            "signals": signals,
        }
        anomalies: list[str] = []
        for item in members:
            anomalies.extend(item.evidence.get("anomalies") or [])
            for field in (
                "complaint_level2",
                "complaint_level3",
                "source_level_code",
                "severity_rank",
                "absence_reason_detail",
                "is_late",
                "is_early",
                "is_fake_early",
            ):
                if field in item.evidence and field not in evidence:
                    evidence[field] = item.evidence[field]
        if anomalies:
            evidence["anomalies"] = list(dict.fromkeys(anomalies))
        why = " ".join(dict.fromkeys(item.why for item in members))
        scope_key = f"lesson:{row.lesson_id}"
        dedupe_key = f"{canonical_rule}:{row.teacher_id}:{row.lesson_id}"
        if output_type == "TEACHER_TASK":
            if task_code == "P-FB-COMPLAINT":
                complaint_key = hashlib.sha256(
                    normalize_text(row.complaint_l3).encode("utf-8")
                ).hexdigest()[:16]
                output_dedupe_key = (
                    f"{canonical_rule}:{row.teacher_id}:complaint:{complaint_key}"
                )
            else:
                output_dedupe_key = f"{canonical_rule}:{row.teacher_id}"
        else:
            output_dedupe_key = dedupe_key
        result.append(
            _OutputSpec(
                rule_code=canonical_rule,
                domain=members[0].domain,
                output_type=output_type,
                title=members[0].title,
                priority=_priority_min(item.priority for item in members),
                why=why,
                evidence=evidence,
                teacher_id=row.teacher_id,
                lesson_id=row.lesson_id,
                source_record_id=source_record_id,
                complaint_rule_id=complaint_rule_id,
                scope_key=scope_key,
                dedupe_key=dedupe_key,
                output_dedupe_key=output_dedupe_key,
                task_code=task_code,
            )
        )
    return result


def _build_output_specs(
    lessons: Sequence[_LessonRow],
    *,
    lesson_batch_id: str,
    complaint_rules: Mapping[str, ComplaintRule],
    complaint_rule_ids: Mapping[str, str],
) -> tuple[list[_OutputSpec], int, int, int]:
    result: list[_OutputSpec] = []
    blacklists: dict[str, dict[str, _LessonRow]] = defaultdict(dict)
    negative_tags: dict[str, dict[str, list[_LessonRow]]] = defaultdict(
        lambda: defaultdict(list)
    )
    negative_rows_without_labels: dict[str, list[_LessonRow]] = defaultdict(list)
    unmatched_complaints = 0
    for row in lessons:
        source_record_id = _source_record_id(lesson_batch_id, row.row_number)
        complaint_rule_id = complaint_rule_ids.get(normalize_text(row.complaint_l3))
        decisions = evaluate_lesson(
            row.raw_payload,
            complaint_rules=complaint_rules,
        )
        unmatched_complaints += sum(
            item.output_type == "PENDING_DATA" for item in decisions
        )
        result.extend(
            _merge_lesson_decisions(
                row,
                decisions,
                source_record_id=source_record_id,
                complaint_rule_id=complaint_rule_id,
            )
        )
        if row.is_blocked:
            blacklists[row.teacher_id].setdefault(row.student_id, row)
        if row.has_negative_tag:
            if row.negative_tags:
                for label in row.negative_tags:
                    negative_tags[row.teacher_id][label].append(row)
            else:
                negative_rows_without_labels[row.teacher_id].append(row)

    blacklist_count = 0
    for teacher_id, student_rows in blacklists.items():
        if len(student_rows) < 2:
            continue
        blacklist_count += 1
        ordered = sorted(student_rows.values(), key=lambda item: (item.local_start_at, item.row_number))
        threshold_row = ordered[1]
        result.append(
            _OutputSpec(
                rule_code="TR-FB-BLACKLIST",
                domain="USER_FEEDBACK",
                output_type="TEACHER_TASK",
                title="拉黑问题",
                priority="P1",
                why=(
                    f"已有 {len(student_rows)} 名不同学员拉黑该教师，请完成拉黑改善学习任务。"
                ),
                evidence={
                    "distinct_student_count": len(student_rows),
                    "threshold": 2,
                    "threshold_crossing_lesson_id": threshold_row.lesson_id,
                    "lesson_ids": [item.lesson_id for item in ordered],
                },
                teacher_id=teacher_id,
                lesson_id=threshold_row.lesson_id,
                source_record_id=_source_record_id(lesson_batch_id, threshold_row.row_number),
                complaint_rule_id=None,
                scope_key=f"teacher:{teacher_id}:blacklist",
                dedupe_key=f"TR-FB-BLACKLIST:{teacher_id}",
                output_dedupe_key=f"TR-FB-BLACKLIST:{teacher_id}",
                task_code="P-FB-BLACKLIST",
            )
        )

    for teacher_id, labels in negative_tags.items():
        for label, rows in labels.items():
            if len(rows) <= 1:
                continue
            ordered = sorted(
                rows,
                key=lambda item: (item.local_start_at, item.row_number),
            )
            threshold_row = ordered[1]
            label_key = hashlib.sha256(
                normalize_text(label).encode("utf-8")
            ).hexdigest()[:16]
            result.append(
                _OutputSpec(
                    rule_code="TR-FB-NEGATIVE-REPEAT",
                    domain="USER_FEEDBACK",
                    output_type="TEACHER_TASK",
                    title=f"差评-{label}问题",
                    priority="P1",
                    why=(
                        f"同一差评标签“{label}”已在该教师的 {len(rows)} 节差评课中出现，"
                        "请完成对应学习任务。"
                    ),
                    evidence={
                        "negative_feedback_label": label,
                        "negative_review_lesson_count": len(rows),
                        "aggregate_hit_count": len(rows),
                        "threshold": 2,
                        "threshold_crossing_lesson_id": threshold_row.lesson_id,
                        "lesson_ids": [item.lesson_id for item in ordered],
                    },
                    teacher_id=teacher_id,
                    lesson_id=threshold_row.lesson_id,
                    source_record_id=_source_record_id(
                        lesson_batch_id, threshold_row.row_number
                    ),
                    complaint_rule_id=None,
                    scope_key=f"teacher:{teacher_id}:negative-label:{label_key}",
                    dedupe_key=f"TR-FB-NEGATIVE-REPEAT:{teacher_id}:{label_key}",
                    output_dedupe_key=(
                        f"TR-FB-NEGATIVE-REPEAT:{teacher_id}:{label_key}"
                    ),
                    task_code="P-FB-NEGATIVE",
                )
            )

    negative_pending_count = 0
    for teacher_id, rows in negative_rows_without_labels.items():
        if len(rows) <= 1:
            continue
        negative_pending_count += 1
        ordered = sorted(rows, key=lambda item: (item.local_start_at, item.row_number))
        threshold_row = ordered[1]
        result.append(
            _OutputSpec(
                rule_code="TR-FB-NEGATIVE-TAG-MISSING",
                domain="USER_FEEDBACK",
                output_type="PENDING_DATA",
                title="差评任务待补标签名称",
                priority="P1",
                why=(
                    "差评标签为 1，但评价详情没有可拆分的评价标签；同一教师已重复出现，"
                    "因此暂不向教师发布任务。"
                ),
                evidence={
                    "negative_tag_flag_count": len(rows),
                    "threshold": 2,
                    "missing_field": "评价详情.评价标签",
                    "lesson_ids": [item.lesson_id for item in ordered],
                },
                teacher_id=teacher_id,
                lesson_id=threshold_row.lesson_id,
                source_record_id=_source_record_id(lesson_batch_id, threshold_row.row_number),
                complaint_rule_id=None,
                scope_key=f"teacher:{teacher_id}:negative-tag-missing",
                dedupe_key=f"TR-FB-NEGATIVE-TAG-MISSING:{teacher_id}",
                output_dedupe_key=f"TR-FB-NEGATIVE-TAG-MISSING:{teacher_id}",
                task_code=None,
            )
        )
    return result, blacklist_count, negative_pending_count, unmatched_complaints


def _make_batch(
    *,
    batch_id: str,
    source_kind: str,
    source_system: str,
    source: Path,
    source_sha256: str,
    source_sheet: str,
    snapshot_label: str,
    column_count: int,
    row_count: int,
    header: Sequence[str],
    imported_at: datetime,
    payload: dict[str, Any],
) -> DataImportBatchRecord:
    return DataImportBatchRecord(
        batch_id=batch_id,
        source_kind=source_kind,
        sync_mode="MANUAL_BASELINE",
        source_system=source_system,
        source_filename=source.name,
        source_uri=str(source.resolve()),
        source_sha256=source_sha256,
        source_sheet=source_sheet,
        snapshot_label=snapshot_label,
        data_mode="REAL",
        column_count=column_count,
        row_count=row_count,
        header=list(header),
        status="COMPLETED",
        imported_at=imported_at,
        payload=payload,
        created_at=imported_at,
        updated_at=imported_at,
    )


def _add_complaint_batch(
    session: Session,
    *,
    source: Path,
    source_sha256: str,
    batch_id: str,
    rules: Sequence[_ComplaintRow],
    skipped_rows: Sequence[int],
    raw_rows: Sequence[dict[str, Any]],
    imported_at: datetime,
) -> int:
    session.add(
        _make_batch(
            batch_id=batch_id,
            source_kind="COMPLAINT_RULES",
            source_system=COMPLAINT_SOURCE_SYSTEM,
            source=source,
            source_sha256=source_sha256,
            source_sheet=COMPLAINT_SOURCE_SHEET,
            snapshot_label="CURRENT_COMPLAINT_LEVELS_20260722",
            column_count=6,
            row_count=len(raw_rows),
            header=EXPECTED_COMPLAINT_HEADERS,
            imported_at=imported_at,
            payload={
                "source_region": COMPLAINT_SOURCE_REGION,
                "rule_count": len(rules),
                "unmapped_source_rows": list(skipped_rows),
                "mapping_key": "三级分类精确匹配",
            },
        )
    )
    # These models intentionally do not expose ORM relationships. Flush the
    # parent first so PostgreSQL can enforce the FK without relying on unit-of-
    # work relationship ordering; the surrounding transaction remains atomic.
    session.flush()
    for offset, raw in enumerate(raw_rows, start=3):
        session.add(
            SourceRecord(
                source_record_id=_source_record_id(batch_id, offset),
                batch_id=batch_id,
                source_sheet=COMPLAINT_SOURCE_SHEET,
                source_row_number=offset,
                business_key=f"complaint-source-row:{offset}",
                teacher_id=None,
                lesson_id=None,
                occurred_at=None,
                row_sha256=hashlib.sha256(_canonical_json(raw).encode("utf-8")).hexdigest(),
                raw_payload=raw,
                created_at=imported_at,
            )
        )
    for item in rules:
        assert item.category_l3 is not None
        session.add(
            ComplaintCategoryRuleRecord(
                rule_id=_complaint_rule_id(batch_id, item.row_number),
                batch_id=batch_id,
                source_sheet=COMPLAINT_SOURCE_SHEET,
                source_row_number=item.row_number,
                category_l1=item.category_l1,
                category_l2=item.category_l2,
                category_l3=item.category_l3,
                category_l3_normalized=normalize_text(item.category_l3),
                source_level=item.source_level,
                normalized_level=item.normalized_level,
                severity_rank=item.severity_rank,
                default_route=item.default_route,
                learning_title=item.learning_title,
                learning_url=item.learning_url,
                raw_payload={
                    "source_row": item.raw_payload,
                    "effective_category_l1": item.category_l1,
                    "effective_category_l2": item.category_l2,
                },
                created_at=imported_at,
            )
        )
    session.flush()
    return len(raw_rows)


def _student_hash(student_id: str) -> str:
    return hashlib.sha256(f"TIT-STUDENT:{student_id}".encode("utf-8")).hexdigest()


def _lesson_fact_kwargs(
    row: _LessonRow,
    *,
    batch_id: str,
    teacher: TeacherRecord,
    complaint_rule: _ComplaintRow | None,
    complaint_rule_id: str | None,
    imported_at: datetime,
) -> dict[str, Any]:
    kwargs: dict[str, Any] = {
        "lesson_id": row.lesson_id,
        "source_appoint_id": row.lesson_id,
        "camp_enrollment_id": teacher.camp_enrollment_id,
        "teacher_id": row.teacher_id,
        "scheduled_start_at": None,
        "scheduled_end_at": None,
        "lesson_lifecycle_status": row.lifecycle_status,
        "lesson_local_date": row.local_date,
        "lesson_local_time": row.local_time,
        "student_id_hash": _student_hash(row.student_id),
        "is_peak": row.is_peak,
        "is_late": row.is_late,
        "is_early": row.is_early,
        "is_false_early_leave": row.is_false_early_leave,
        "negative_score": row.negative_score,
        "feedback_detail": row.feedback_detail,
        "negative_tag_values": list(row.negative_tags),
        "absence_reason_detail": row.absence_reason_detail,
        "complaint_category_l1": row.complaint_l1,
        "complaint_category_l2": row.complaint_l2,
        "complaint_category_l3": row.complaint_l3,
        "complaint_source_level": complaint_rule.source_level if complaint_rule else None,
        "complaint_level_rank": complaint_rule.severity_rank if complaint_rule else None,
        "complaint_route": complaint_rule.default_route if complaint_rule else None,
        "complaint_rule_id": complaint_rule_id,
        "is_blocked": row.is_blocked,
        "is_favorited": row.is_favorited,
        "positive_tag_value": None,
        "is_rebooked": row.is_rebooked,
        "is_camera_off": row.is_camera_off,
        "is_cpu_usage_high": row.is_cpu_usage_high,
        "is_network_delay_high": row.is_network_delay_high,
        "source_batch_id": batch_id,
        "source_record_id": _source_record_id(batch_id, row.row_number),
        "valid_for_scoring": False,
        "evidence_status": "OBSERVED_REAL_SOURCE",
        "data_mode": "REAL",
        "payload": {
            "source_batch_id": batch_id,
            "source_record_id": _source_record_id(batch_id, row.row_number),
            "source_row_number": row.row_number,
            "lesson_local_start_at": row.local_start_at.isoformat(),
            "score_note": "本批课程事实仅用于个性化任务触发；课程积分仍按已确认教师统计口径结算。",
        },
        "created_at": imported_at,
        "updated_at": imported_at,
    }
    # These typed flags were added after the original lesson model. Keeping
    # the guard lets an older disposable SQLite schema fail only at migration
    # time while preserving the raw source row in all supported schemas.
    if hasattr(LessonFactRecord, "has_negative_feedback_tag"):
        kwargs["has_negative_feedback_tag"] = row.has_negative_tag
    if hasattr(LessonFactRecord, "has_positive_feedback_tag"):
        kwargs["has_positive_feedback_tag"] = row.has_positive_tag
    return kwargs


def _add_lesson_batch(
    session: Session,
    *,
    source: Path,
    source_sha256: str,
    sheet_name: str,
    batch_id: str,
    lessons: Sequence[_LessonRow],
    teachers: Mapping[str, TeacherRecord],
    complaint_rows: Mapping[str, _ComplaintRow],
    complaint_rule_ids: Mapping[str, str],
    imported_at: datetime,
    supersedes_batch_ids: Sequence[str] = (),
) -> int:
    min_date = min(item.local_date for item in lessons)
    max_date = max(item.local_date for item in lessons)
    batch = _make_batch(
        batch_id=batch_id,
        source_kind="LESSON_BASELINE",
        source_system=LESSON_SOURCE_SYSTEM,
        source=source,
        source_sha256=source_sha256,
        source_sheet=sheet_name,
        snapshot_label=f"LESSONS_{min_date.isoformat()}_{max_date.isoformat()}",
        column_count=len(EXPECTED_LESSON_HEADERS),
        row_count=len(lessons),
        header=EXPECTED_LESSON_HEADERS,
        imported_at=imported_at,
        payload={
            "date_range": [min_date.isoformat(), max_date.isoformat()],
            "teacher_count": len({item.teacher_id for item in lessons}),
            "student_identifier_storage": "RAW_RESTRICTED_AND_TYPED_SHA256",
            "trigger_rule_version": TRIGGER_RULE_VERSION,
            "is_current": True,
            "supersedes_batch_ids": list(supersedes_batch_ids),
        },
    )
    session.add(batch)
    session.flush()
    for row in lessons:
        source_id = _source_record_id(batch_id, row.row_number)
        session.add(
            SourceRecord(
                source_record_id=source_id,
                batch_id=batch_id,
                source_sheet=sheet_name,
                source_row_number=row.row_number,
                business_key=row.lesson_id,
                teacher_id=row.teacher_id,
                lesson_id=row.lesson_id,
                # The workbook provides no source timezone. Local time stays
                # in typed lesson columns and raw_payload; no UTC instant is invented.
                occurred_at=None,
                row_sha256=hashlib.sha256(
                    _canonical_json(row.raw_payload).encode("utf-8")
                ).hexdigest(),
                raw_payload=row.raw_payload,
                created_at=imported_at,
            )
        )
    session.flush()
    existing_facts = {
        item.lesson_id: item
        for chunk in _chunks(sorted(item.lesson_id for item in lessons))
        for item in session.scalars(
            select(LessonFactRecord).where(LessonFactRecord.lesson_id.in_(chunk))
        ).all()
    }
    for row in lessons:
        normalized_l3 = normalize_text(row.complaint_l3)
        complaint_rule = complaint_rows.get(normalized_l3)
        complaint_rule_id = complaint_rule_ids.get(normalized_l3)
        values = _lesson_fact_kwargs(
            row,
            batch_id=batch_id,
            teacher=teachers[row.teacher_id],
            complaint_rule=complaint_rule,
            complaint_rule_id=complaint_rule_id,
            imported_at=imported_at,
        )
        existing = existing_facts.get(row.lesson_id)
        if existing is None:
            session.add(LessonFactRecord(**values))
            continue
        for field, value in values.items():
            if field not in {"lesson_id", "created_at"}:
                setattr(existing, field, value)
    session.flush()
    return len(lessons)


def _prepare_lesson_baseline_replacement(
    session: Session,
    *,
    new_batch_id: str,
    new_lesson_ids: set[str],
    replaced_at: datetime,
) -> list[str]:
    """Retire the current lesson projection without erasing task or raw history.

    New lesson facts replace the current read projection. Existing outputs are
    reconciled separately: still-valid outputs are reused, unconsumed stale
    outputs are cancelled, and consumed outputs remain immutable history.
    """

    lesson_batch_ids = [
        item.batch_id
        for item in session.scalars(
            select(DataImportBatchRecord).where(
                DataImportBatchRecord.source_kind == "LESSON_BASELINE",
                DataImportBatchRecord.source_system == LESSON_SOURCE_SYSTEM,
            )
        ).all()
    ]
    current_lessons = (
        session.scalars(
            select(LessonFactRecord).where(
                LessonFactRecord.source_batch_id.in_(lesson_batch_ids)
            )
        ).all()
        if lesson_batch_ids
        else []
    )
    if not current_lessons:
        return []
    current_lesson_ids = {item.lesson_id for item in current_lessons}
    missing_from_replacement = sorted(current_lesson_ids - new_lesson_ids)
    if missing_from_replacement:
        raise LessonImportValidationError(
            "replacement baseline must contain every currently projected lesson; "
            f"missing_count={len(missing_from_replacement)}, "
            f"sample={missing_from_replacement[:10]!r}"
        )

    old_batch_ids = sorted(
        {
            item.source_batch_id
            for item in current_lessons
            if item.source_batch_id is not None
        }
    )
    for old_batch in session.scalars(
        select(DataImportBatchRecord).where(
            DataImportBatchRecord.batch_id.in_(old_batch_ids)
        )
    ).all():
        old_batch.payload = {
            **(old_batch.payload if isinstance(old_batch.payload, dict) else {}),
            "is_current": False,
            "superseded_by_batch_id": new_batch_id,
            "superseded_at": replaced_at.isoformat(),
        }
        old_batch.updated_at = replaced_at

    session.flush()
    return old_batch_ids


def _materialize_outputs(
    session: Session,
    *,
    specs: Sequence[_OutputSpec],
    templates: Mapping[str, TaskTemplateRecord],
    teachers: Mapping[str, TeacherRecord],
    materialized_at: datetime,
    reconcile_existing: bool = False,
) -> dict[str, int]:
    counts = Counter(
        task_assignments_created=0,
        ops_cases_created=0,
        notifications_created=0,
        pending_data_matches_created=0,
        trigger_matches_created=0,
    )
    existing_matches = {
        item.dedupe_key: item
        for item in session.scalars(select(PersonalizedTriggerMatchRecord)).all()
    }
    task_specs = [item for item in specs if item.output_type == "TEACHER_TASK"]
    task_groups: dict[str, list[_OutputSpec]] = defaultdict(list)
    for item in task_specs:
        task_groups[item.output_dedupe_key].append(item)
    existing_tasks = {
        item.dedupe_key: item
        for chunk in _chunks(sorted(task_groups))
        for item in session.scalars(
            select(TaskAssignmentRecord).where(TaskAssignmentRecord.dedupe_key.in_(chunk))
        ).all()
    }

    task_outputs: dict[str, TaskAssignmentRecord] = {}
    for output_dedupe_key, members in task_groups.items():
        assignment = existing_tasks.get(output_dedupe_key)
        first = members[0]
        if not first.task_code:
            raise LessonImportValidationError(
                f"{first.rule_code}: teacher task has no task_code"
            )
        if any(
            item.task_code != first.task_code or item.title != first.title
            for item in members
        ):
            raise LessonImportValidationError(
                f"task aggregation conflict for {output_dedupe_key}"
            )
        lesson_ids = list(
            dict.fromkeys(
                lesson_id
                for item in members
                for lesson_id in (
                    item.evidence.get("lesson_ids")
                    if isinstance(item.evidence.get("lesson_ids"), list)
                    else [item.lesson_id]
                )
                if lesson_id
            )
        )
        hit_count = sum(
            int(item.evidence.get("aggregate_hit_count") or 1)
            for item in members
        )
        evidence = {
            "hit_count": hit_count,
            "lesson_ids": lesson_ids,
            "matched_rule_codes": list(
                dict.fromkeys(item.rule_code for item in members)
            ),
            "signal_samples": [
                {
                    "lesson_id": item.lesson_id,
                    "why": item.why,
                    "evidence": item.evidence,
                }
                for item in members[:20]
            ],
            "sample_limit": 20,
            "evidence_is_complete_for_lesson_ids": True,
        }
        lesson_sample = "、".join(lesson_ids[:10])
        sample_suffix = "……" if len(lesson_ids) > 10 else ""
        why = (
            f"共命中 {hit_count} 节课程（课程 ID：{lesson_sample}{sample_suffix}）。"
            f"{first.why}"
        )
        if assignment is None:
            template = templates[first.task_code]
            template_payload = template.payload if isinstance(template.payload, dict) else {}
            due_hours = int(
                (template_payload.get("due_rule") or {}).get("hours") or 72
            )
            output_id = _stable_id("TASK", output_dedupe_key)
            teacher = teachers[first.teacher_id]
            assignment = TaskAssignmentRecord(
                assignment_id=output_id,
                teacher_id=first.teacher_id,
                task_code=first.task_code,
                template_version_id=template.row_id,
                task_kind="PERSONALIZED_IMPROVEMENT",
                creator_system="TRIGGER_CENTER",
                status="ASSIGNED",
                priority=_priority_min(item.priority for item in members),
                why=why,
                display_title=first.title,
                evidence_snapshot=evidence,
                due_at=materialized_at + timedelta(hours=due_hours),
                timezone_used=teacher.timezone,
                timezone_source="TEACHER_PROFILE",
                timezone_verified_at=materialized_at,
                status_reason_code=None,
                source_mode="DERIVED_REAL",
                dedupe_key=output_dedupe_key,
                created_by="TRIGGER_CENTER",
                updated_by="TRIGGER_CENTER",
                row_version=1,
                assigned_at=materialized_at,
                status_changed_at=materialized_at,
                completed_at=None,
                created_at=materialized_at,
                updated_at=materialized_at,
            )
            session.add(assignment)
            existing_tasks[output_dedupe_key] = assignment
            counts["task_assignments_created"] += 1
        task_outputs[output_dedupe_key] = assignment

    for spec in specs:
        output_id: str | None = None
        match_status = (
            "PENDING_DATA" if spec.output_type == "PENDING_DATA" else "MATERIALIZED"
        )
        if spec.output_type == "TEACHER_TASK":
            assignment = task_outputs[spec.output_dedupe_key]
            output_id = assignment.assignment_id
        elif spec.output_type == "OPS_CASE":
            output_id = _stable_id("CASE", spec.output_dedupe_key)
            if session.get(OpsCaseRecord, output_id) is None:
                session.add(
                    OpsCaseRecord(
                        case_id=output_id,
                        case_type="SEVERE_COMPLAINT",
                        teacher_id=spec.teacher_id,
                        task_id=None,
                        priority=spec.priority,
                        status="OPEN",
                        source_reason=spec.rule_code,
                        external_action_status="OPS_REVIEW_REQUIRED",
                        created_at=materialized_at,
                        payload={
                            "title": spec.title,
                            "summary": spec.why,
                            "recommended_action": "核实投诉事实并按现行处罚规则处理。",
                            "evidence": spec.evidence,
                            "source_mode": "DERIVED_REAL",
                            "trigger_rule_version": TRIGGER_RULE_VERSION,
                        },
                        updated_at=materialized_at,
                    )
                )
                counts["ops_cases_created"] += 1
        elif spec.output_type == "NOTIFICATION":
            output_id = _stable_id("NOTIF", spec.output_dedupe_key)
            if session.get(NotificationRecord, output_id) is None:
                session.add(
                    NotificationRecord(
                        notification_id=output_id,
                        task_id=None,
                        source_ref=spec.output_dedupe_key,
                        teacher_id=spec.teacher_id,
                        channel="WEBAPP_INBOX",
                        priority=spec.priority,
                        status="STORED",
                        requested_at=materialized_at,
                        stored_at=materialized_at,
                        read_at=None,
                        clicked_at=None,
                        response_due_at=None,
                        failure_reason=None,
                        payload={
                            "title": spec.title,
                            "body": spec.why,
                            "evidence": spec.evidence,
                            "source_mode": "DERIVED_REAL",
                            "trigger_rule_version": TRIGGER_RULE_VERSION,
                        },
                    )
                )
                counts["notifications_created"] += 1
        elif spec.output_type == "PENDING_DATA":
            if spec.dedupe_key not in existing_matches:
                counts["pending_data_matches_created"] += 1
        else:
            raise LessonImportValidationError(
                f"unsupported personalized output type: {spec.output_type}"
            )

        existing_match = existing_matches.get(spec.dedupe_key)
        match_values = {
            "trigger_code": spec.rule_code,
            "rule_version": TRIGGER_RULE_VERSION,
            "teacher_id": spec.teacher_id,
            "lesson_id": spec.lesson_id,
            "source_record_id": spec.source_record_id,
            "complaint_rule_id": spec.complaint_rule_id,
            "scope_key": spec.scope_key,
            "output_type": spec.output_type,
            "output_title": spec.title,
            "output_id": output_id,
            "match_status": match_status,
            "evidence_snapshot": {
                **spec.evidence,
                "why": spec.why,
                "source_mode": "DERIVED_REAL",
            },
            "materialized_at": (
                materialized_at if match_status == "MATERIALIZED" else None
            ),
            "updated_at": materialized_at,
        }
        if existing_match is not None:
            for field, value in match_values.items():
                setattr(existing_match, field, value)
        else:
            session.add(
                PersonalizedTriggerMatchRecord(
                    trigger_match_id=_stable_id("TM", spec.dedupe_key),
                    dedupe_key=spec.dedupe_key,
                    **match_values,
                    matched_at=materialized_at,
                    created_at=materialized_at,
                )
            )
            counts["trigger_matches_created"] += 1

    if reconcile_existing:
        desired_match_keys = {item.dedupe_key for item in specs}
        for dedupe_key, match in existing_matches.items():
            if dedupe_key in desired_match_keys:
                continue
            match.match_status = "SUPPRESSED"
            match.evidence_snapshot = {
                **(
                    match.evidence_snapshot
                    if isinstance(match.evidence_snapshot, dict)
                    else {}
                ),
                "superseded_by_rule_version": TRIGGER_RULE_VERSION,
                "superseded_at": materialized_at.isoformat(),
            }
            match.updated_at = materialized_at

        session.flush()
        active_output_ids = set(
            session.scalars(
                select(PersonalizedTriggerMatchRecord.output_id).where(
                    PersonalizedTriggerMatchRecord.match_status != "SUPPRESSED",
                    PersonalizedTriggerMatchRecord.output_id.is_not(None),
                )
            ).all()
        )
        for assignment in session.scalars(
            select(TaskAssignmentRecord).where(
                TaskAssignmentRecord.task_kind == "PERSONALIZED_IMPROVEMENT",
                TaskAssignmentRecord.creator_system == "TRIGGER_CENTER",
            )
        ).all():
            if (
                assignment.assignment_id not in active_output_ids
                and assignment.status == "ASSIGNED"
            ):
                assignment.status = "CANCELLED"
                assignment.status_reason_code = "SOURCE_EVIDENCE_SUPERSEDED"
                assignment.status_changed_at = materialized_at
                assignment.updated_by = "TRIGGER_CENTER_BASELINE_REPLACEMENT"
        for notification in session.scalars(select(NotificationRecord)).all():
            if (
                notification.notification_id not in active_output_ids
                and notification.status == "STORED"
            ):
                notification.status = "CANCELLED"
        for case in session.scalars(
            select(OpsCaseRecord).where(OpsCaseRecord.source_reason.is_not(None))
        ).all():
            if case.case_id not in active_output_ids and case.status == "OPEN":
                case.status = "CANCELLED"
                case.external_action_status = "SOURCE_EVIDENCE_SUPERSEDED"
                case.updated_at = materialized_at
    return dict(counts)


def _result_from_stored_report(
    batch: DataImportBatchRecord,
    *,
    dry_run: bool,
) -> LessonBaselineImportResult:
    report = dict((batch.payload or {}).get("import_report") or {})
    if not report:
        raise LessonImportValidationError(
            f"completed lesson batch {batch.batch_id} has no immutable import report"
        )
    report["dry_run"] = dry_run
    report["idempotent_reimport"] = True
    return LessonBaselineImportResult(**report)


def import_lesson_baseline(
    lesson_source: str | Path,
    complaint_source: str | Path,
    *,
    bind: Engine | None = None,
    expected_lesson_row_count: int | None = EXPECTED_LESSON_ROW_COUNT,
    dry_run: bool = False,
    replace_current: bool = False,
) -> LessonBaselineImportResult:
    """Atomically import the real lesson baseline, rules and derived outputs.

    The function performs a full file and database preflight before adding a
    single row. It is content-idempotent by source SHA-256 + sheet, never
    creates placeholder teachers, and keeps student identifiers out of typed
    operator-facing lesson facts.
    """

    lesson_path = Path(lesson_source).expanduser().resolve()
    complaint_path = Path(complaint_source).expanduser().resolve()
    lesson_sheet, lesson_sha, lessons = _read_lesson_workbook(
        lesson_path,
        expected_row_count=expected_lesson_row_count,
    )
    complaint_sha, complaint_rows, skipped_complaint_rows, raw_complaint_rows = (
        _read_complaint_workbook(complaint_path)
    )
    lesson_batch_id = f"LESSON-{lesson_sha[:24]}"
    complaint_batch_id = f"COMPLAINT-{complaint_sha[:24]}"
    complaint_rules, complaint_rule_ids = _complaint_rule_context(
        complaint_rows,
        complaint_batch_id=complaint_batch_id,
    )
    complaint_rows_by_l3 = {
        normalize_text(item.category_l3): item
        for item in complaint_rows
        if item.category_l3 is not None
    }
    selected_engine = bind or default_engine
    session = Session(selected_engine, expire_on_commit=False)
    try:
        existing_lesson_batch = session.scalar(
            select(DataImportBatchRecord).where(
                DataImportBatchRecord.source_sha256 == lesson_sha,
                DataImportBatchRecord.source_sheet == lesson_sheet,
            )
        )
        if existing_lesson_batch is not None:
            if existing_lesson_batch.status != "COMPLETED":
                raise LessonImportValidationError(
                    f"lesson batch {existing_lesson_batch.batch_id} is not completed"
                )
            return _result_from_stored_report(existing_lesson_batch, dry_run=dry_run)

        teachers = _teacher_map(session, (item.teacher_id for item in lessons))
        templates = _template_map(session)
        imported_at = _now()
        existing_lessons = _existing_values(
            session,
            LessonFactRecord.lesson_id,
            [item.lesson_id for item in lessons],
        )
        supersedes_batch_ids: list[str] = []
        if existing_lessons and not replace_current:
            raise LessonImportValidationError(
                "lesson preflight failed: lesson ids already exist in another batch; "
                f"conflict_count={len(existing_lessons)}"
            )
        if existing_lessons:
            supersedes_batch_ids = _prepare_lesson_baseline_replacement(
                session,
                new_batch_id=lesson_batch_id,
                new_lesson_ids={item.lesson_id for item in lessons},
                replaced_at=imported_at,
            )
        output_specs, blacklist_count, negative_pending_count, unmatched_complaints = (
            _build_output_specs(
                lessons,
                lesson_batch_id=lesson_batch_id,
                complaint_rules=complaint_rules,
                complaint_rule_ids=complaint_rule_ids,
            )
        )
        trigger_counts = dict(Counter(item.rule_code for item in output_specs))
        proposed_counts = Counter(item.output_type for item in output_specs)
        proposed_task_assignment_count = len(
            {
                item.output_dedupe_key
                for item in output_specs
                if item.output_type == "TEACHER_TASK"
            }
        )
        existing_complaint_batch = session.scalar(
            select(DataImportBatchRecord).where(
                DataImportBatchRecord.source_sha256 == complaint_sha,
                DataImportBatchRecord.source_sheet == COMPLAINT_SOURCE_SHEET,
            )
        )
        complaint_source_records_to_create = 0 if existing_complaint_batch else len(raw_complaint_rows)
        base_result = {
            "lesson_batch_id": lesson_batch_id,
            "complaint_batch_id": complaint_batch_id,
            "lesson_source_sha256": lesson_sha,
            "complaint_source_sha256": complaint_sha,
            "lesson_sheet": lesson_sheet,
            "lesson_row_count": len(lessons),
            "complaint_source_row_count": len(raw_complaint_rows),
            "complaint_rule_count": len(complaint_rows),
            "complaint_unmapped_source_rows": list(skipped_complaint_rows),
            "source_records_created": len(lessons) + complaint_source_records_to_create,
            "lesson_facts_created": len(lessons),
            "task_assignments_created": proposed_task_assignment_count,
            "ops_cases_created": proposed_counts["OPS_CASE"],
            "notifications_created": proposed_counts["NOTIFICATION"],
            "pending_data_matches_created": proposed_counts["PENDING_DATA"],
            "trigger_matches_created": len(output_specs),
            "blacklist_tasks_created": blacklist_count,
            "negative_tag_pending_teachers": negative_pending_count,
            "unmatched_complaint_lessons": unmatched_complaints,
            "trigger_counts": trigger_counts,
            "dry_run": dry_run,
            "idempotent_reimport": False,
        }
        if dry_run:
            session.rollback()
            return LessonBaselineImportResult(**base_result)

        if existing_complaint_batch is None:
            _add_complaint_batch(
                session,
                source=complaint_path,
                source_sha256=complaint_sha,
                batch_id=complaint_batch_id,
                rules=complaint_rows,
                skipped_rows=skipped_complaint_rows,
                raw_rows=raw_complaint_rows,
                imported_at=imported_at,
            )
        elif existing_complaint_batch.status != "COMPLETED":
            raise LessonImportValidationError(
                f"complaint batch {existing_complaint_batch.batch_id} is not completed"
            )
        elif existing_complaint_batch.batch_id != complaint_batch_id:
            raise LessonImportValidationError(
                "complaint content id conflicts with its deterministic batch id"
            )

        _add_lesson_batch(
            session,
            source=lesson_path,
            source_sha256=lesson_sha,
            sheet_name=lesson_sheet,
            batch_id=lesson_batch_id,
            lessons=lessons,
            teachers=teachers,
            complaint_rows=complaint_rows_by_l3,
            complaint_rule_ids=complaint_rule_ids,
            imported_at=imported_at,
            supersedes_batch_ids=supersedes_batch_ids,
        )
        materialized_counts = _materialize_outputs(
            session,
            specs=output_specs,
            templates=templates,
            teachers=teachers,
            materialized_at=imported_at,
            reconcile_existing=replace_current,
        )
        base_result.update(materialized_counts)
        lesson_batch = session.get(DataImportBatchRecord, lesson_batch_id)
        if lesson_batch is None:
            # Pending instances are not visible through get() until flush on
            # every SQLAlchemy backend, so locate it in the identity map path.
            session.flush()
            lesson_batch = session.get(DataImportBatchRecord, lesson_batch_id)
        assert lesson_batch is not None
        lesson_batch.payload = {
            **lesson_batch.payload,
            "import_report": base_result,
        }
        session.commit()
        return LessonBaselineImportResult(**base_result)
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()
