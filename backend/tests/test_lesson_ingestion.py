from __future__ import annotations

from datetime import datetime

from openpyxl import Workbook, load_workbook
import pytest
from sqlalchemy import func, select

from app.database import engine, session_scope
from app.db_models import (
    DataImportBatchRecord,
    LessonFactRecord,
    NotificationRecord,
    OpsCaseRecord,
    PersonalizedTriggerMatchRecord,
    SourceRecord,
    TaskAssignmentRecord,
    TeacherRecord,
)
from app.lesson_ingestion import (
    COMPLAINT_SOURCE_SHEET,
    EXPECTED_COMPLAINT_HEADERS,
    EXPECTED_LESSON_HEADERS,
    LessonImportValidationError,
    import_lesson_baseline,
)


def _complaint_workbook(path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = COMPLAINT_SOURCE_SHEET
    sheet.append(["现行版本"])
    sheet.append(list(EXPECTED_COMPLAINT_HEADERS))
    for row_number in range(3, 46):
        if row_number == 35:
            sheet.append(["关于老师", "老师不适当的索要好评", None, "P1", None, None])
            continue
        level2 = "教学技巧问题"
        level3 = f"分类-{row_number}"
        level = "P4"
        if row_number == 3:
            level3 = "一般投诉样例"
        elif row_number == 4:
            level2 = "严重投诉"
            level3 = "严重投诉样例"
            level = "P0"
        elif row_number == 5:
            level2 = "网络设备问题"
            level3 = "网络卡顿"
        sheet.append(["关于老师", level2, level3, level, "Learning", "https://example.test"])
    workbook.save(path)


def _lesson_workbook(path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LESSON_BASELINE"
    sheet.append(list(EXPECTED_LESSON_HEADERS))

    def row(
        lesson_id: int,
        student_id: int,
        *,
        memo: bool = False,
        late: bool = False,
        negative_tag: bool = False,
        blocked: bool = False,
        complaint_l2=None,
        complaint_l3=None,
        camera_off: bool = False,
    ):
        payload = {
            "课程id": lesson_id,
            "上课日期": datetime(2026, 4, 22),
            "上课时间": datetime(2026, 4, 22, 9, lesson_id % 60),
            "是否高峰": 1,
            "老师id": 90001,
            "学员id": student_id,
            "课程状态": "end",
            "缺席原因明细": "Unfilled Lesson Memo" if memo else None,
            "迟到": int(late),
            "早退": 0,
            "差评分": 1 if negative_tag else None,
            "差评标签": int(negative_tag),
            "投诉一级分类": "关于老师" if complaint_l2 else None,
            "投诉二级分类": complaint_l2,
            "投诉三级分类": complaint_l3,
            "是否拉黑": int(blocked),
            "收藏": 0,
            "好评标签": 0,
            "评价详情": "缺少互动, 缺乏热情" if negative_tag else None,
            "是否复约": 0,
            "未开摄像头": int(camera_off),
            "cpu占用过高": 0,
            "网络延迟过高": 0,
            "假早退": 0,
        }
        return [payload[header] for header in EXPECTED_LESSON_HEADERS]

    sheet.append(row(1001, 70001, memo=True, late=True, negative_tag=True, blocked=True))
    sheet.append(row(1002, 70002, memo=True, negative_tag=True, blocked=True))
    sheet.append(row(1003, 70003, late=True))
    sheet.append(
        row(
            1004,
            70004,
            complaint_l2="网络设备问题",
            complaint_l3="网络卡顿",
            camera_off=True,
        )
    )
    sheet.append(
        row(1005, 70005, complaint_l2="严重投诉", complaint_l3="严重投诉样例")
    )
    sheet.append(
        row(1006, 70006, complaint_l2="教学技巧问题", complaint_l3="一般投诉样例")
    )
    workbook.save(path)


def test_real_lesson_import_is_atomic_aggregated_and_idempotent(tmp_path) -> None:
    lessons = tmp_path / "lessons.xlsx"
    complaints = tmp_path / "complaints.xlsx"
    _lesson_workbook(lessons)
    _complaint_workbook(complaints)
    with session_scope(engine) as session:
        session.add(
            TeacherRecord(
                teacher_id="90001",
                camp_enrollment_id="CAMP-90001",
                name="Import Test Teacher",
                country="PH",
                timezone="Asia/Manila",
                camp_day=5,
                graduation_state="IN_PROGRESS",
                total_score=0,
                graduation_threshold=80,
                data_mode="MIXED",
                source_batch_id=None,
                source_snapshot_label="TEST",
                payload={"teacher_id": "90001"},
            )
        )

    dry_run = import_lesson_baseline(
        lessons,
        complaints,
        bind=engine,
        expected_lesson_row_count=6,
        dry_run=True,
    )
    assert dry_run.lesson_facts_created == 6
    assert dry_run.task_assignments_created == 6
    assert dry_run.notifications_created == 1
    assert dry_run.ops_cases_created == 1
    assert dry_run.pending_data_matches_created == 0
    with session_scope(engine) as session:
        assert session.scalar(select(func.count()).select_from(LessonFactRecord)) == 1

    result = import_lesson_baseline(
        lessons,
        complaints,
        bind=engine,
        expected_lesson_row_count=6,
    )
    assert result.source_records_created == 49
    assert result.lesson_facts_created == 6
    assert result.task_assignments_created == 6
    assert result.trigger_matches_created == 9

    with session_scope(engine) as session:
        assert session.scalar(
            select(func.count()).select_from(SourceRecord).where(
                SourceRecord.batch_id == result.lesson_batch_id
            )
        ) == 6
        assignments = session.scalars(
            select(TaskAssignmentRecord).where(
                TaskAssignmentRecord.creator_system == "TRIGGER_CENTER"
            )
        ).all()
        assert {item.task_code for item in assignments} == {
            "P-REL-MEMO",
            "P-REL-ATTENDANCE",
            "P-FB-COMPLAINT",
            "P-FB-BLACKLIST",
            "P-FB-NEGATIVE",
        }
        memo = next(item for item in assignments if item.task_code == "P-REL-MEMO")
        assert memo.evidence_snapshot["hit_count"] == 2
        assert memo.evidence_snapshot["lesson_ids"] == ["1001", "1002"]
        memo_matches = session.scalars(
            select(PersonalizedTriggerMatchRecord).where(
                PersonalizedTriggerMatchRecord.trigger_code == "TR-REL-LESSON-MEMO"
            )
        ).all()
        assert len(memo_matches) == 2
        assert {item.output_id for item in memo_matches} == {memo.assignment_id}
        quality = session.scalar(
            select(PersonalizedTriggerMatchRecord).where(
                PersonalizedTriggerMatchRecord.trigger_code == "TR-QUALITY-LESSON"
            )
        )
        assert quality is not None
        assert set(quality.evidence_snapshot["matched_rule_codes"]) == {
            "TR-QUALITY-COMPLAINT",
            "TR-QUALITY-IN-CLASS",
        }
        assert session.scalar(select(func.count()).select_from(NotificationRecord)) == 1
        assert session.scalar(select(func.count()).select_from(OpsCaseRecord)) == 1
        negative_tasks = [
            item for item in assignments if item.task_code == "P-FB-NEGATIVE"
        ]
        assert {item.display_title for item in negative_tasks} == {
            "差评-缺少互动问题",
            "差评-缺乏热情问题",
        }
        assert all(item.evidence_snapshot["hit_count"] == 2 for item in negative_tasks)

    repeated = import_lesson_baseline(
        lessons,
        complaints,
        bind=engine,
        expected_lesson_row_count=6,
    )
    assert repeated.idempotent_reimport is True
    with session_scope(engine) as session:
        assert session.scalar(
            select(func.count()).select_from(TaskAssignmentRecord).where(
                TaskAssignmentRecord.creator_system == "TRIGGER_CENTER"
            )
        ) == 6


def test_reviewed_superset_can_replace_unconsumed_lesson_projection(tmp_path) -> None:
    lessons_v1 = tmp_path / "lessons-v1.xlsx"
    lessons_v2 = tmp_path / "lessons-v2.xlsx"
    complaints = tmp_path / "complaints.xlsx"
    _lesson_workbook(lessons_v1)
    _lesson_workbook(lessons_v2)
    _complaint_workbook(complaints)
    workbook = load_workbook(lessons_v2)
    sheet = workbook.active
    added = list(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True)))
    added[EXPECTED_LESSON_HEADERS.index("课程id")] = 1007
    added[EXPECTED_LESSON_HEADERS.index("学员id")] = 70007
    added[EXPECTED_LESSON_HEADERS.index("差评标签")] = 0
    added[EXPECTED_LESSON_HEADERS.index("评价详情")] = None
    sheet.append(added)
    workbook.save(lessons_v2)

    with session_scope(engine) as session:
        session.add(
            TeacherRecord(
                teacher_id="90001",
                camp_enrollment_id="CAMP-90001",
                name="Replacement Test Teacher",
                country="PH",
                timezone="Asia/Manila",
                camp_day=5,
                graduation_state="IN_PROGRESS",
                total_score=0,
                graduation_threshold=80,
                data_mode="MIXED",
                source_batch_id=None,
                source_snapshot_label="TEST",
                payload={"teacher_id": "90001"},
            )
        )

    first = import_lesson_baseline(
        lessons_v1,
        complaints,
        bind=engine,
        expected_lesson_row_count=6,
    )
    replacement = import_lesson_baseline(
        lessons_v2,
        complaints,
        bind=engine,
        expected_lesson_row_count=7,
        replace_current=True,
    )

    assert replacement.lesson_facts_created == 7
    with session_scope(engine) as session:
        assert session.scalar(
            select(func.count()).select_from(LessonFactRecord).where(
                LessonFactRecord.source_batch_id == replacement.lesson_batch_id
            )
        ) == 7
        old_batch = session.get(DataImportBatchRecord, first.lesson_batch_id)
        new_batch = session.get(DataImportBatchRecord, replacement.lesson_batch_id)
        assert old_batch is not None
        assert old_batch.payload["is_current"] is False
        assert old_batch.payload["superseded_by_batch_id"] == replacement.lesson_batch_id
        assert new_batch is not None
        assert new_batch.payload["is_current"] is True
        assert new_batch.payload["supersedes_batch_ids"] == [first.lesson_batch_id]
        assert session.scalar(
            select(func.count()).select_from(SourceRecord).where(
                SourceRecord.batch_id == first.lesson_batch_id
            )
        ) == 6


def test_missing_teacher_fails_preflight_without_partial_batch(tmp_path) -> None:
    lessons = tmp_path / "lessons.xlsx"
    complaints = tmp_path / "complaints.xlsx"
    _lesson_workbook(lessons)
    _complaint_workbook(complaints)

    with pytest.raises(LessonImportValidationError, match="teacher foreign keys are missing"):
        import_lesson_baseline(
            lessons,
            complaints,
            bind=engine,
            expected_lesson_row_count=6,
        )

    with session_scope(engine) as session:
        assert session.scalar(select(func.count()).select_from(SourceRecord)) == 0
